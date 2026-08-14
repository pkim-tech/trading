"""Consolidated liquidity-screen candidate summary -- for each ticker, THREE
candidate rows (best safe node / best unsafe node / 5min best possible node,
see CANDIDATE_LABELS below), each with core alpha, raw return, ANNUALIZED
excess return (CAGR-based, fair across tickers with different cached-history
lengths), 5-min real-fill accuracy, worst neighbor (cliff-safety), liquidity,
and drought/add-on overlay results (only shown when that exact row's params
match a registered candidate_nodes entry -- overlay backtests are only ever
run against ONE specific node's params, not all three candidate types, so
showing the same overlay numbers on every row would misrepresent them as
validated for configs they were never tested against). Built 2026-08-08
after repeatedly rebuilding pieces of this table by hand in conversation --
consolidates locate_best_node.py's winner pick, top_safe_nodes.py's
neighbor-check logic, candidate_5min_report.py's 3-way candidate split,
annualized_alpha_report.py's CAGR calc, verify_fill_resolution_accuracy.py's
5-min replay, and candidate_overlay_results into one script. The canonical
"everything at a glance" candidate report -- keep adding to this rather than
spinning up another standalone comparison script, per the user's explicit
call 2026-08-08 (later). Row structure (3 rows/ticker, not 1) is also the
user's explicit call, same session -- "i would look at best unsafe first to
just make sure we're not missing anything."

Usage:
  .venv/bin/python scripts/candidate_summary_report.py TNA URTY SQQQ ...
  .venv/bin/python scripts/candidate_summary_report.py --all-swept
  .venv/bin/python scripts/candidate_summary_report.py TNA --skip-5min  # faster, skip yfinance calls
  .venv/bin/python scripts/candidate_summary_report.py TNA --xlsx report  # output/report.xlsx, 2 sheets
"""
import argparse
import csv
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from top_safe_nodes import CLIFF_RADIUS
from annualized_alpha_report import calendar_days, cagr
from locate_best_node import resolve_version
from verify_fill_resolution_accuracy import fill_accuracy_for_node
from candidate_5min_report import find_candidates
from run_overlay_shim import (
    run_for_node as run_overlay_for_node, ensure_candidate_nodes_table, ensure_table as ensure_overlay_table,
)
from datetime import datetime as _datetime

DB_PATH = "cache/research/trading_universe.db"
ROBUST_ALPHA_SQL = ("MIN(alpha_vs_spy, COALESCE(alpha_vs_spy_pessimistic, alpha_vs_spy), "
                     "COALESCE(alpha_vs_spy_certain, alpha_vs_spy))")

# Relabels find_candidates()'s internal keys to the user's requested wording
# 2026-08-08 (later) -- kept as a separate map (not renamed at the source in
# candidate_5min_report.py) since that script's own terminal output is aimed
# at a different, more technical audience/context.
CANDIDATE_LABELS = {
    'cliff-safe (current convention)': 'best safe node',
    'CAGR-first (possible-resolution, cliff-checked on the same metric)': 'best CAGR-safe node',
    'best robust_alpha (ignoring cliff-safety)': 'best unsafe node',
    'best possible (raw alpha_vs_spy)': '5min best possible',
    'best certain (alpha_vs_spy_certain, no-guessing resolution)': 'best certain',
}
CANDIDATE_TYPE_ORDER = ['best safe node', 'best CAGR-safe node', 'best unsafe node', '5min best possible',
                         'best certain']

# Single source of truth for what each output column means -- used for both
# the xlsx glossary sheet and (imported directly) the Streamlit Candidates
# page's glossary expander, so the two never drift apart.
COLUMN_DEFS = {
    "ticker": "The symbol.",
    "candidate_type": "Which of the 4 candidate-selection methods produced this row: 'best safe node' "
                       "(robust_alpha=MIN(possible,pessimistic,certain), required to pass the CLIFF_RADIUS=3 "
                       "neighbor-safety check -- this project's real selection convention), 'best CAGR-safe "
                       "node' (same cliff-safety check but ranked on alpha_vs_spy/CAGR-equivalent instead of "
                       "robust_alpha -- your actual real-world selection preference), 'best unsafe node' "
                       "(top robust_alpha row regardless of whether it clears the safety check -- look here "
                       "first to sanity-check the safe pick isn't missing something), or '5min best possible' "
                       "(top raw 'possible'-resolution alpha_vs_spy row, ignoring robustness entirely -- named "
                       "for the 5-min fill-accuracy finding that 'possible' is empirically the most accurate "
                       "single resolution, see docs/research_log.md's 2026-08-08 entries). Rows are dropped, "
                       "not duplicated, when two of the three methods pick the identical node.",
    "strategy": "Which strategy class this ticker's rows use (TrailingBothZScoreBreakout = trailing-buy entry, "
                "TrailingExitZScoreBreakout = market-buy entry) -- always the strategy of the ticker's single "
                "best robust_alpha row across all strategies; all 3 candidate rows for a ticker share it.",
    "core_alpha_pct": "robust_alpha for this row's node: MIN(possible, pessimistic, certain) fill-resolution "
                       "alpha vs SPY, over the ticker's full cached-data window. This project's standard "
                       "selection metric -- WITH SPY subtracted (see abs_return_pct for the raw number).",
    "abs_return_pct": "This row's node's own RAW compounded return (SPY NOT subtracted) over the same window. "
                       "Read this next to core_alpha_pct when weighing a 'safe but small' pick against a "
                       "'ridiculous but risky' one -- core_alpha_pct/worst_neighbor_pct are SPY-adjusted, this "
                       "one isn't.",
    "years": "Real calendar span of this ticker's cached hourly data (days between the earliest and latest "
             "cached bar, /365.25). Tickers vary a lot here (e.g. SOXL ~3.0y vs SPCL ~0.31y) -- this is why "
             "core_alpha_pct/abs_return_pct alone aren't fairly comparable across tickers.",
    "trades": "Number of completed trades in this row's node's backtest.",
    "ann_excess_pct": "CAGR-based excess return over SPY (SPY-adjusted, like core_alpha_pct), annualized over "
                       "the full calendar window (years column) -- fixes the cross-ticker horizon-mismatch "
                       "problem above. A large value backed by a short 'years'/low 'trades' is weak evidence "
                       "(e.g. SPCL's ~3000%+ off 0.31y/3 trades is an annualization artifact, not a real "
                       "signal) -- always read this next to years/trades, never alone.",
    "fillacc_possible_win_pct": "Of this row's node's real trailing-buy entry signals in the last ~58 days "
                                 "(yfinance's 5-min history cap), the % where the 'possible' fill resolution "
                                 "(the kernel's default, optimistic-but-unmodified bounce-fill assumption) was "
                                 "closest to the REAL 5-minute-bar fill price, vs the 'pessimistic'/'certain' "
                                 "alternatives. Blank for TrailingExitZScoreBreakout rows (market-buy entry has "
                                 "no bounce-fill resolution to check) or if trail_buy_pct=0.",
    "fillacc_possible_mean_err_pct": "Mean absolute price error (%) of the 'possible' resolution vs the real "
                                      "5-min fill, across those same signals. Lower is better/more trustworthy. "
                                      "Same blank-condition as fillacc_possible_win_pct.",
    "fillacc_n": "How many real signals the fill-accuracy check above is actually based on -- often single "
                 "digits in a 58-day window, so treat a 100% win rate on n=1-2 with real caution.",
    "worst_neighbor_pct": "Cliff-safety check: the worst robust_alpha found among nearby take_profit/stop_loss "
                           "grid values (CLIFF_RADIUS=3 steps) AND max_hold_hours +/-7 around this row's node "
                           "params, holding every other axis fixed -- deliberately matches "
                           "top_safe_nodes.best_safe_node()'s own tolerance (found 2026-08-08: a wider +/-24 "
                           "hold tolerance, prune_backtest_cache.py's own convention, could make a node "
                           "certified 'safe' by best_safe_node() show CLIFF here, purely from a hold-time "
                           "window inconsistency, not a real disagreement about safety). Negative means a "
                           "nearby parameter nudge would have lost money -- see "
                           "docs/cliff_safety_query_checklist.md.",
    "status": "CLIFF (worst_neighbor_pct < 0, fragile to small parameter changes) or SAFE (worst_neighbor_pct "
              ">= 0), computed fresh for THIS row's own node -- so 'best unsafe node'/'5min best possible' rows "
              "can (and often do) show CLIFF even though 'best safe node' always shows SAFE by construction.",
    "addon_n": "Number of backtested add-on-leg trades for THIS exact row's own node params -- computed on "
               "demand (run_overlay_shim.py) the first time a candidate is seen, then reused from "
               "candidate_overlay_results on later runs. All 3 candidate rows for a ticker get their own real "
               "run against their own params, not a shared/repeated number. Blank only if the node has <2 "
               "real trades to evaluate an overlay against, or --skip-overlay was passed.",
    "addon_compounded_pct": "Compounded return of just the add-on overlay's own trades (not combined with core). "
                             "Same blank-condition as addon_n.",
    "addon_win_rate_pct": "% of add-on trades that were profitable. Same blank-condition as addon_n.",
    "drought_n": "Number of backtested drought-overlay trades, same on-demand-per-row-node computation and "
                 "blank-condition as addon_n.",
    "drought_compounded_pct": "Compounded return of just the drought overlay's own trades (not combined with core).",
    "drought_win_rate_pct": "% of drought trades that were profitable.",
    "x_addon_pct": "NAIVE estimate of core+add-on combined: (1+core_return)*(1+addon_return)-1. Add-on capital "
                   "runs CONCURRENTLY with an open core position (not sequentially), so this OVERSTATES the "
                   "real combined effect -- v5_stacked_backtest.py's parallel-return model is the rigorous "
                   "version. Treat this column as a rough upper bound, not a real number.",
    "x_drought_pct": "NAIVE estimate of core+drought combined, same (1+core)*(1+overlay)-1 formula. Drought "
                      "fills core's own idle-time gaps SEQUENTIALLY, so this approximation is more defensible "
                      "than x_addon_pct, but still not the rigorous stacked model.",
    "liquidity_dollars_per_day": "avg_vol_10d * last_price * 0.01 from the tickers table -- this project's "
                                  "standard real dollar-liquidity estimate (confirmed against CLAUDE.md's cited "
                                  "figures, e.g. AGQ ~$2.02M, HIBL ~$89.9K). Ticker-level, same across all 3 "
                                  "candidate rows for a ticker. Supposed to be the FIRST-pass filter before "
                                  "spending validation effort on a candidate (see the 2026-08-07 'liquidity was "
                                  "never the limiting filter' finding) -- a great alpha number on an illiquid "
                                  "name is untradeable regardless of the rest of this row.",
}


def best_node_strategy(conn, ticker, version="v5"):
    """Which strategy has this ticker's single best robust_alpha row, across
    ALL strategies -- used to scope the 3-way candidate split to one strategy
    (mixing TrailingBoth/TrailingExit params in a neighbor search wouldn't be
    meaningful, same convention top_safe_nodes.py already uses). Also returns
    spy_bh (ticker-level, identical across every row for this ticker/version,
    so fetched once here rather than per candidate)."""
    c = conn.cursor()
    c.execute(f"""
        SELECT strategy, spy_bh FROM backtest_cache
        WHERE ticker=? AND version=? AND trades>0
        ORDER BY {ROBUST_ALPHA_SQL} DESC LIMIT 1
    """, (ticker, version))
    return c.fetchone()


def load_ticker_df(conn, ticker, version, strategy):
    import pandas as pd
    df = pd.read_sql("""
        SELECT ticker, COALESCE(take_profit, arm_sell_pct) AS take_profit, stop_loss, max_hold_hours, window,
               z_score_threshold, trail_buy_pct, trail_sell_pct, entry_timing,
               alpha_vs_spy, alpha_vs_spy_pessimistic, alpha_vs_spy_certain,
               strategy_return, trades, win_rate, sweep_run_id
        FROM backtest_cache
        WHERE ticker=? AND version=? AND strategy=? AND trades>0
    """, conn, params=(ticker, version, strategy))
    pess = df["alpha_vs_spy_pessimistic"].fillna(df["alpha_vs_spy"])
    cert = df["alpha_vs_spy_certain"].fillna(df["alpha_vs_spy"])
    df["robust_alpha"] = pd.concat([df["alpha_vs_spy"], pess, cert], axis=1).min(axis=1)
    return df


def annualized_excess(ticker, strategy_return, spy_bh_full_window):
    """CAGR-based excess return over the ticker's full cached-data calendar
    span -- see annualized_alpha_report.py's docstring for why raw alpha_vs_spy
    isn't cross-ticker comparable (different tickers have very different
    amounts of cached history) and why this annualizes over the full window
    rather than invested-only time (user's explicit call, 2026-08-08)."""
    days = calendar_days(ticker)
    strat_cagr = cagr(strategy_return, days)
    spy_cagr = cagr(spy_bh_full_window, days)
    if strat_cagr is None or spy_cagr is None:
        return days, None
    return days, strat_cagr - spy_cagr


def fill_accuracy_summary(ticker, strategy, window, z, trail_buy_pct, hold):
    """(possible_win_rate_pct, possible_mean_abs_err_pct, n) from the 5-min
    real-fill replay, or None if this node's entry mechanism has no bounce-
    fill resolution to check (TrailingExitZScoreBreakout's market-buy entry,
    or a TrailingBoth row with trail_buy_pct=0)."""
    if strategy != "TrailingBothZScoreBreakout" or not trail_buy_pct:
        return None
    df = fill_accuracy_for_node(ticker, window, z, trail_buy_pct, hold)
    if df.empty:
        return None
    diff_cols = ['possible_diff_pct', 'pessimistic_diff_pct', 'certain_diff_pct']
    abs_diffs = df[diff_cols].abs()
    closest = abs_diffs.idxmin(axis=1)
    win_rate = (closest == 'possible_diff_pct').mean() * 100
    mean_abs_err = df['possible_diff_pct'].dropna().abs().mean()
    return win_rate, mean_abs_err, len(df)


def worst_neighbor(conn, ticker, version, strategy, window, z, entry_timing,
                    sl, tp, hold, tb, ts, metric="robust_alpha"):
    """Worst `metric` among the CLIFF_RADIUS-step neighborhood on
    take_profit/stop_loss (index-based nearest-distinct-values), holding
    every other axis fixed. max_hold_hours tolerance intentionally set to
    +/-7 (not prune_backtest_cache.py's own +/-24) to MATCH
    top_safe_nodes.best_safe_node()'s tolerance -- found 2026-08-08 that a
    node top_safe_nodes.py certified 'safe' (+2.7% worst neighbor at +/-7)
    came back CLIFF here at +/-24 (-9.8%) for TNA, purely because +/-24
    happened to span this ticker's ENTIRE swept hold-time range (6 values,
    7 apart) rather than just nearby ones. Since this report's 'best safe
    node'/'best CAGR-safe node' labels come directly from best_safe_node()'s
    own check, this function must use the SAME tolerance AND the SAME
    metric that check used, or the label and Status column can visibly
    disagree -- found 2026-08-11 for 'best CAGR-safe node': this function
    defaulted to robust_alpha unconditionally, so a node genuinely verified
    safe under alpha_vs_spy (its own selection metric) printed a
    contradictory CLIFF status computed against a DIFFERENT metric than the
    one that vetted it. prune_backtest_cache.py's +/-24 is left untouched --
    that's a real, separate, deliberate convention for island selection, not
    something to change as a side effect of this report."""
    metric_sql = ROBUST_ALPHA_SQL if metric == "robust_alpha" else metric
    c = conn.cursor()

    def nearest_values(col_sql, center):
        c.execute(f"""
            SELECT DISTINCT {col_sql} FROM backtest_cache
            WHERE ticker=? AND version=? AND strategy=? AND window=? AND z_score_threshold=?
                  AND entry_timing=? AND trail_buy_pct=? AND trail_sell_pct=?
        """, (ticker, version, strategy, window, z, entry_timing, tb, ts))
        vals = sorted(set(r[0] for r in c.fetchall() if r[0] is not None))
        if center not in vals:
            return {center}
        idx = vals.index(center)
        return set(vals[max(0, idx - CLIFF_RADIUS):min(len(vals), idx + CLIFF_RADIUS + 1)])

    tp_keep = list(nearest_values("COALESCE(take_profit, arm_sell_pct)", tp))
    sl_keep = list(nearest_values("stop_loss", sl))
    tp_ph = ",".join("?" * len(tp_keep))
    sl_ph = ",".join("?" * len(sl_keep))
    c.execute(f"""
        SELECT MIN({metric_sql}) FROM backtest_cache
        WHERE ticker=? AND version=? AND strategy=? AND window=? AND z_score_threshold=?
              AND entry_timing=? AND trail_buy_pct=? AND trail_sell_pct=?
              AND COALESCE(take_profit, arm_sell_pct) IN ({tp_ph})
              AND stop_loss IN ({sl_ph})
              AND ABS(max_hold_hours - ?) <= 7 AND trades > 0
    """, [ticker, version, strategy, window, z, entry_timing, tb, ts] + tp_keep + sl_keep + [hold])
    return c.fetchone()[0]


def compounded(rets):
    prod = 1.0
    for r in rets:
        prod *= (1 + r)
    return (prod - 1) * 100


def liquidity_dollars_per_day(conn, ticker):
    """Real dollar liquidity, this project's standard formula (see
    campaign_comparison_table.py) -- avg_vol_10d*last_price*0.01, confirmed
    against CLAUDE.md's cited figures (AGQ ~$2.02M/day, HIBL ~$89.9K/day)."""
    c = conn.cursor()
    c.execute("SELECT avg_vol_10d * last_price * 0.01 FROM tickers WHERE symbol=?", (ticker,))
    row = c.fetchone()
    return row[0] if row else None


def overlay_summary_for_node(conn, ticker, strategy, version, mechanism, node):
    """Overlay backtests (candidate_overlay_results) are only ever run
    against ONE specific node's exact params (via candidate_nodes/
    run_overlay_shim.py), not against all 3 candidate types -- so this
    matches on the full param tuple, not just ticker, and returns None
    (blank) for a candidate row that was never actually run through the
    overlay shim, rather than showing another row's numbers as if they
    applied here too."""
    c = conn.cursor()
    # Real gap found in paired review 2026-08-09 (candidate_full_review.py):
    # run_overlay_shim.py's INSERT has no dedup and can be re-run for the
    # same node -- 22+ real candidate_node_ids have 2-5 duplicate
    # run_timestamps for the same mechanism, silently inflating n/compounded
    # here. Scoped to the latest run_timestamp per (candidate_node_id,
    # mechanism), matching the same fix in candidate_full_review.overlay_robustness.
    c.execute("""
        SELECT cor.ret FROM candidate_overlay_results cor
        JOIN candidate_nodes cn ON cn.id = cor.candidate_node_id
        WHERE cn.ticker=? AND cor.mechanism=? AND cn.strategy=? AND cn.version=?
              AND cn.window=? AND cn.z=? AND cn.fixed_sl=? AND cn.arm_pct=?
              AND cn.trail_buy_pct=? AND cn.trail_sell_pct=? AND cn.max_hold_hours=? AND cn.entry_timing=?
              AND cor.run_timestamp = (
                  SELECT MAX(cor2.run_timestamp) FROM candidate_overlay_results cor2
                  WHERE cor2.candidate_node_id = cor.candidate_node_id AND cor2.mechanism = cor.mechanism
              )
    """, (ticker, mechanism, strategy, version, node['window'], node['z'], node['sl'], node['arm_pct'],
          node['trail_buy_pct'], node['trail_sell_pct'], node['hold'], node['entry_timing']))
    rets = [r[0] for r in c.fetchall()]
    if not rets:
        return None
    wr = sum(1 for r in rets if r > 0) / len(rets) * 100
    return len(rets), compounded(rets), wr


def ensure_overlay_for_node(conn, ticker, strategy, version, node, confirm_days=10):
    """Computes drought/addon overlay results on demand for THIS exact node
    if they're not already in candidate_overlay_results -- added 2026-08-08
    (later) per the user's explicit call: 'all three candidates should get
    the same overlay treatment', not just whichever one happened to already
    be registered from an earlier locate_best_node.py/run_overlay_shim.py
    run. Only computes the missing mechanism(s); commits immediately so a
    later run's overlay_summary_for_node lookup (and other tools reading
    candidate_overlay_results) see it too."""
    missing = {m for m in ("drought", "addon")
               if overlay_summary_for_node(conn, ticker, strategy, version, m, node) is None}
    if not missing:
        return
    shim_node = {
        'ticker': ticker, 'strategy': strategy, 'version': version,
        'window': node['window'], 'z': node['z'], 'fixed_sl': node['sl'],
        'arm_pct': node['arm_pct'], 'trail_buy_pct': node['trail_buy_pct'],
        'trail_sell_pct': node['trail_sell_pct'], 'max_hold_hours': node['hold'],
        'entry_timing': node['entry_timing'],
        'robust_alpha': node['robust_alpha'], 'trades': node['trades'],
    }
    rows = run_overlay_for_node(conn, ticker, shim_node, confirm_days, mechanisms=missing)
    if not rows:
        return
    run_ts = _datetime.now().isoformat(timespec="seconds")
    conn.executemany("""
        INSERT INTO candidate_overlay_results
            (run_timestamp, mechanism, ticker, candidate_node_id,
             confirm_days, entry_time, exit_time, exit_reason, ret)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [(run_ts, r["mechanism"], r["ticker"], r["candidate_node_id"],
           r["confirm_days"], r["entry_time"], r["exit_time"], r["exit_reason"], r["ret"])
          for r in rows])
    conn.commit()


def _row_to_record(row):
    """Converts one internal row tuple to a {column_name: value} dict keyed
    exactly like COLUMN_DEFS, so CSV/xlsx/terminal output and the glossary
    all agree on column names in one place."""
    ticker = row[0]
    if row[1] is None:
        return {"ticker": ticker}
    (ticker, candidate_type, strategy, ralpha, sret, years, trades, ann_excess, fill_acc, wn, cliff,
     addon, drought, addon_mult, drought_mult, liquidity) = row
    return {
        "ticker": ticker, "candidate_type": candidate_type, "strategy": strategy,
        "core_alpha_pct": ralpha, "abs_return_pct": sret,
        "years": years, "trades": trades, "ann_excess_pct": ann_excess,
        "fillacc_possible_win_pct": fill_acc[0] if fill_acc else None,
        "fillacc_possible_mean_err_pct": fill_acc[1] if fill_acc else None,
        "fillacc_n": fill_acc[2] if fill_acc else None,
        "worst_neighbor_pct": wn, "status": cliff,
        "addon_n": addon[0] if addon else None,
        "addon_compounded_pct": addon[1] if addon else None,
        "addon_win_rate_pct": addon[2] if addon else None,
        "drought_n": drought[0] if drought else None,
        "drought_compounded_pct": drought[1] if drought else None,
        "drought_win_rate_pct": drought[2] if drought else None,
        "x_addon_pct": addon_mult, "x_drought_pct": drought_mult,
        "liquidity_dollars_per_day": liquidity,
    }


def _write_csv(name, rows):
    out_path = Path("output") / (name if name.endswith(".csv") else f"{name}.csv")
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(COLUMN_DEFS.keys()))
        w.writeheader()
        for row in rows:
            w.writerow(_row_to_record(row))
    print(f"Wrote {out_path} ({len(rows)} rows)")


def _write_xlsx(name, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path("output") / (name if name.endswith(".xlsx") else f"{name}.xlsx")
    out_path.parent.mkdir(exist_ok=True)

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Candidates"

    cols = list(COLUMN_DEFS.keys())
    data_ws.append(cols)
    for cell in data_ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        rec = _row_to_record(row)
        data_ws.append([rec.get(c) for c in cols])
    data_ws.freeze_panes = "A2"
    for i, col in enumerate(cols, start=1):
        data_ws.column_dimensions[get_column_letter(i)].width = max(12, min(len(col) + 2, 28))

    def_ws = wb.create_sheet("Column Definitions")
    def_ws.append(["Column", "Definition"])
    for cell in def_ws[1]:
        cell.font = Font(bold=True)
    for col, definition in COLUMN_DEFS.items():
        def_ws.append([col, definition])
        def_ws.cell(row=def_ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    def_ws.column_dimensions["A"].width = 32
    def_ws.column_dimensions["B"].width = 110

    wb.save(out_path)
    print(f"Wrote {out_path} ({len(rows)} rows, 2 sheets)")


def build_rows_for_ticker(conn, ticker, version, min_alpha, skip_5min, skip_overlay=False):
    """Returns a list of internal row tuples, one per candidate type (up to
    3, deduped -- see find_candidates()), or [(ticker, None)] if there's no
    data/no candidates at all."""
    best = best_node_strategy(conn, ticker, version)
    if best is None:
        return [(ticker, None)]
    strategy, spy_bh = best

    df_t = load_ticker_df(conn, ticker, version, strategy)
    candidates = find_candidates(df_t, min_alpha)
    if not candidates:
        return [(ticker, None)]

    # Only 'CAGR-first' gets a non-default metric here -- it's the one new
    # type whose own selection check (inside best_safe_node) used
    # alpha_vs_spy, not robust_alpha, so its Status/WorstNb must be
    # recomputed on that SAME metric or it visibly contradicts its own
    # label (found 2026-08-11, see worst_neighbor's docstring). The other
    # 4 types keep their existing, unchanged robust_alpha-based Status --
    # deliberately not widened to '5min best possible'/'best certain' too,
    # which have their own separate, longstanding, deliberate "show the
    # robust_alpha view even though a different metric picked this node"
    # framing that isn't part of this fix's scope.
    WN_METRIC_BY_LABEL = {
        'CAGR-first (possible-resolution, cliff-checked on the same metric)': 'alpha_vs_spy',
    }

    rows = []
    for raw_label, node in candidates.items():
        label = CANDIDATE_LABELS.get(raw_label, raw_label)
        wn = worst_neighbor(conn, ticker, version, strategy, node['window'], node['z'],
                             node['entry_timing'], node['sl'], node['arm_pct'], node['hold'],
                             node['trail_buy_pct'], node['trail_sell_pct'],
                             metric=WN_METRIC_BY_LABEL.get(raw_label, "robust_alpha"))
        cliff = "CLIFF" if (wn is not None and wn < 0) else ("SAFE" if wn is not None else "?")
        days, ann_excess = annualized_excess(ticker, node['return'], spy_bh)
        years = round(days / 365.25, 2) if days else None
        fill_acc = None if skip_5min else fill_accuracy_summary(
            ticker, strategy, node['window'], node['z'], node['trail_buy_pct'], node['hold'])
        if not skip_overlay:
            ensure_overlay_for_node(conn, ticker, strategy, version, node)
        addon = overlay_summary_for_node(conn, ticker, strategy, version, "addon", node)
        drought = overlay_summary_for_node(conn, ticker, strategy, version, "drought", node)
        # NAIVE multiplicative combination -- (1 + core) * (1 + overlay) - 1.
        # This is an APPROXIMATION, not real stacked-model math: drought fills
        # core's own time gaps sequentially (so multiplying compounded
        # multipliers is roughly defensible), but add-on runs CONCURRENTLY
        # with an open core position (parallel capital, not sequential), so
        # multiplying it against core here overstates/misrepresents the real
        # combined effect -- v5_stacked_backtest.py's proper parallel-return
        # model is the rigorous version of this, not this quick estimate.
        core_mult = 1 + node['return'] / 100
        addon_mult = (core_mult * (1 + addon[1] / 100) - 1) * 100 if addon else None
        drought_mult = (core_mult * (1 + drought[1] / 100) - 1) * 100 if drought else None
        liquidity = liquidity_dollars_per_day(conn, ticker)
        rows.append((ticker, label, strategy, node['robust_alpha'], node['return'], years, node['trades'],
                     ann_excess, fill_acc, wn, cliff, addon, drought, addon_mult, drought_mult, liquidity))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tickers", nargs="*")
    ap.add_argument("--version", default=None,
                     help="force a single version for every ticker (old behavior). Default: auto-resolve "
                          "per ticker via resolve_version() -- v5.1 when the ticker has it, else v5.")
    ap.add_argument("--db", default=DB_PATH)
    ap.add_argument("--min-alpha", type=float, default=200,
                     help="Alpha floor for the 'best safe node' cliff-safety search (default 200%%, matching "
                          "top_safe_nodes.py's convention). 'best unsafe node'/'5min best possible' are always "
                          "shown regardless of this floor.")
    ap.add_argument("--skip-5min", action="store_true",
                     help="skip the 5-min fill-accuracy replay (saves a yfinance call per ticker)")
    ap.add_argument("--skip-overlay", action="store_true",
                     help="skip computing drought/addon overlay for candidates that don't have it yet "
                          "(faster, but addon/drought columns stay blank for un-registered candidates)")
    ap.add_argument("--csv", default=None, help="write output/<name>.csv instead of the wide terminal table")
    ap.add_argument("--xlsx", default=None,
                     help="write output/<name>.xlsx (Candidates sheet + Column Definitions glossary sheet) "
                          "instead of the wide terminal table")
    args = ap.parse_args()

    conn = sqlite3.connect(args.db)
    ensure_candidate_nodes_table(conn)
    ensure_overlay_table(conn)
    tickers = args.tickers
    if not tickers:
        c = conn.cursor()
        c.execute("SELECT DISTINCT ticker FROM candidate_nodes")
        tickers = [r[0] for r in c.fetchall()]

    rows = []
    for ticker in tickers:
        version = args.version or resolve_version(conn, ticker)
        rows.extend(build_rows_for_ticker(conn, ticker, version, args.min_alpha, args.skip_5min,
                                           args.skip_overlay))

    conn.close()

    if args.xlsx:
        _write_xlsx(args.xlsx, rows)
        return
    if args.csv:
        _write_csv(args.csv, rows)
        return

    hdr = "%-8s %-20s %12s %9s %9s %6s %6s %10s %14s %9s %6s %20s %20s %12s %12s" % (
        "Ticker", "Candidate", "Liquidity$/d", "CoreA%", "AbsRet%", "Years", "Trades", "AnnExcess%",
        "FillAcc(win%,err%)", "WorstNb%", "Status", "Addon(n,comp%,WR%)", "Drought(n,comp%,WR%)",
        "x Addon%", "x Drought%")
    print(hdr)
    print("(x columns are a NAIVE multiplicative estimate, not real stacked-model math -- see docstring)")
    print("(AnnExcess%/CoreA% are SPY-adjusted; AbsRet% is not -- see COLUMN_DEFS. Years/Trades sit next to "
          "AnnExcess%: a big number backed by a short history / few trades is weaker evidence.)")
    print("(Addon/Drought columns are blank unless THIS row's exact params match a registered candidate_nodes "
          "entry that was actually run through the overlay shim -- not repeated across a ticker's 3 rows.)")

    ticker_best = {}
    for row in rows:
        if row[1] is None:
            continue
        ticker_best[row[0]] = max(ticker_best.get(row[0], -1e9), row[3])

    def sort_key(row):
        ticker = row[0]
        best = ticker_best.get(ticker, -1e9) if row[1] is not None else -1e9
        type_rank = CANDIDATE_TYPE_ORDER.index(row[1]) if row[1] in CANDIDATE_TYPE_ORDER else 99
        return (-best, ticker, type_rank)

    for row in sorted(rows, key=sort_key):
        rec = _row_to_record(row)
        if rec.get("core_alpha_pct") is None:
            print(f"{rec['ticker']:8} NO_DATA")
            continue
        wn_str = f"{rec['worst_neighbor_pct']:>9.1f}" if rec['worst_neighbor_pct'] is not None else "      n/a"
        ae_str = f"{rec['ann_excess_pct']:+.1f}" if rec['ann_excess_pct'] is not None else "-"
        fa_str = (f"{rec['fillacc_possible_win_pct']:.0f}%,{rec['fillacc_possible_mean_err_pct']:.2f}%,"
                  f"n={rec['fillacc_n']}") if rec['fillacc_possible_win_pct'] is not None else "-"
        ao_str = (f"{rec['addon_n']},{rec['addon_compounded_pct']:+.2f}%,{rec['addon_win_rate_pct']:.0f}%"
                  if rec['addon_n'] is not None else "-")
        dr_str = (f"{rec['drought_n']},{rec['drought_compounded_pct']:+.2f}%,{rec['drought_win_rate_pct']:.0f}%"
                  if rec['drought_n'] is not None else "-")
        am_str = f"{rec['x_addon_pct']:+.1f}" if rec['x_addon_pct'] is not None else "-"
        dm_str = f"{rec['x_drought_pct']:+.1f}" if rec['x_drought_pct'] is not None else "-"
        liq = rec['liquidity_dollars_per_day']
        liq_str = f"${liq:,.0f}" if liq is not None else "n/a"
        print(f"{rec['ticker']:8} {rec['candidate_type']:<20} {liq_str:>12} {rec['core_alpha_pct']:>9.1f} "
              f"{rec['abs_return_pct']:>9.1f} {rec['years']!s:>6} {rec['trades']:>6} {ae_str:>10} {fa_str:>14} "
              f"{wn_str} {rec['status']:>6} {ao_str:>20} {dr_str:>20} {am_str:>12} {dm_str:>12}")


if __name__ == "__main__":
    main()
