"""Full candidate review in one pass: everything candidate_summary_report.py
already computes (cliff-safety, AnnExcess%, 5-min fill accuracy, liquidity,
addon/drought overlay summary) PLUS several things that previously needed a
separate manual invocation per ticker/mechanism:

1. Raw strategy CAGR (not just AnnExcess% over SPY) -- same cagr() formula
   annualized_alpha_report.py uses, applied to the node's own abs_return_pct
   instead of discarding it after computing the SPY-relative excess.
2. Overlay robustness (docs/overlay_parameter_robustness_process.md steps 1
   and 3, same logic as candidate_overlay_robustness_check.py, reused
   inline instead of requiring a separate per-ticker/per-mechanism run):
   chronological split (are both halves positive?) and single-trade-removal
   (does the sign flip without the single biggest winner?). Also reports
   early/late WIN RATE (not just compounded return) for the same split --
   checklist item 4, applied to drought/addon too, not just core (2026-08-09
   follow-up: found the report was only doing this for core).
3. CORE node out-of-sample walk-forward (5-fold by default, reusing
   walk_forward_check.py's own walk_forward()/summarize() and
   train_test_split_check.py's period_spy_bh() directly against THIS row's
   exact node params via run_backtest_dispatch, not just the old v4/
   TrailingBoth-only best_v4_node() path) -- previously only available as a
   separate per-ticker script run, and only ever checked the v4 winning
   node, never an arbitrary v5 safe/unsafe/possible/certain candidate.
   TrailingExitZScoreBreakout nodes only get the single 'possible'
   resolution (run_backtest_v18 has no pessimistic/certain bounds support,
   unlike TrailingBoth's run_backtest_v110) -- reported fold alpha for
   those rows is possible-only, not the 3-way robust minimum.
4. CORE win-rate stability (checklist item 4) + CORE trade-count fluke check
   (checklist item 8, single-trade-removal on the core node's own trade
   list -- previously only overlay mechanisms got this).
5. Drought included-vs-excluded challenge (docs/overlay_parameter_robustness_process.md
   step 4 -- distinct from step 3's single-trade-removal): does the entry-time
   intraday-vol gate (vol_gate=0.4, the one real validated value, from SOXL's
   confirm_days=3/vol_gate=0.4 signal) actually do real differential selection,
   or do included/excluded windows look similarly good/bad? Addon is excluded
   from this check -- per the taxonomy in overlay_parameter_robustness_process.md,
   addon has no independent parameter to filter on at all (it inherits core's
   own trigger/sizing entirely), so there is no filter to challenge.

Found 2026-08-09 (follow-up session): the first version of this report only
covered 2 of the 4 overlay_parameter_robustness_process.md steps (1 and 3) and
skipped checklist items 4 (framed core-only, not extended to drought/addon)
and 8 (core had no fluke check at all, only overlay mechanisms did) --
despite the module docstring's original "full checklist scope" claim. This
revision closes those gaps.

Built 2026-08-09 after repeatedly being asked to run candidate_summary_report.py,
then annualized_alpha_report.py for raw CAGR, then candidate_overlay_robustness_check.py
per ticker/mechanism as separate steps -- one script, one pass, per the user's
explicit "make 1 script and put it all together" call.

Usage:
  .venv/bin/python scripts/candidate_full_review.py                # all registered candidate_nodes tickers
  .venv/bin/python scripts/candidate_full_review.py SOXL AGQ FAS    # specific tickers
  .venv/bin/python scripts/candidate_full_review.py --skip-5min     # faster, skip yfinance calls
  .venv/bin/python scripts/candidate_full_review.py --csv full_review
"""
import argparse
import csv
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
import yfinance as yf

from annualized_alpha_report import cagr
from candidate_summary_report import (
    DB_PATH, build_rows_for_ticker, _row_to_record, best_node_strategy, load_ticker_df,
    CANDIDATE_LABELS, resolve_version,
)
from candidate_5min_report import best_safe_node, _node_from_row, _node_key
from run_overlay_shim import ensure_candidate_nodes_table, ensure_table as ensure_overlay_table
from locate_best_node import get_or_create_candidate_node, set_pick_comment, get_pick_comment
from walk_forward_check import walk_forward, summarize as summarize_folds
from train_test_split_check import period_spy_bh
from run_optimization_sweep import _load_node_inputs, CACHE_DIR
from backtester import run_backtest_dispatch, run_backtest_v110
from check_stock_splits import check_ticker as _check_splits
from v4_max_drawdown import max_drawdown
from run_optimization_sweep import _summarize_trades
from verify_trailing_sell_resolution import find_hourly_trailing_exits, replay_five_min, FIVE_MIN_LOOKBACK_DAYS
from drought_overlay_test import find_drought_windows, get_trades_and_bars, simulate_overlay
from drought_overlay_sweep import get_ivol_series, _entry_vol_pctile
from sim_bear_market_stress import (
    PROXIES as _BEAR_PROXIES, CRASHES as _BEAR_CRASHES, fetch_underlying as _bear_fetch_underlying,
    synthesize_leveraged as _bear_synthesize, run_strategy_daily as _bear_run_strategy,
    summarize as _bear_summarize,
)
import strategies as _strategies
from datetime import datetime as _dt, timedelta as _timedelta
import pandas as _pd

DEFAULT_VOL_GATE = 0.4  # the one real validated value (SOXL's confirm_days=3/vol_gate=0.4 signal)

STRATEGY_CLASSES = {
    "TrailingBothZScoreBreakout": _strategies.TrailingBothZScoreBreakout,
    "TrailingExitZScoreBreakout": _strategies.TrailingExitZScoreBreakout,
}
DEFAULT_FOLDS = 5

# Single source of truth for column order (both --csv and --xlsx) and what
# each column means (the --xlsx "Column Definitions" glossary sheet) -- same
# pattern as candidate_summary_report.COLUMN_DEFS, extended to this report's
# much larger column set. Added 2026-08-09 after the user expected a real
# glossary sheet by default, not a bare CSV.
COLUMN_DEFS = {
    "ticker": "The symbol.",
    "sector": "Real description/leverage/inverse flag straight from the tickers table.",
    "k1_status": "K-1/UBTI tax-filer status. Only reflects tickers actually confirmed via a real source "
                 "check (issuer tax-document page) -- every other ticker honestly reads 'not checked', never "
                 "guessed from structural resemblance to a confirmed ticker.",
    "k1_tranche": "K1_CONFIRMED / CLEAN_CONFIRMED / ETN_NOT_K1 / NOT_CHECKED bucket of k1_status.",
    "underlier_count": "Number of constituent securities in the underlying index/basket, from real research "
                        "(stockanalysis.com/etf.com/issuer fact sheet) -- NULL if never researched. This "
                        "project's diversification minimum is ~20; below that is a real disqualification risk "
                        "(FNGU=10, BULZ=15, QPUX=13 already disqualified on this basis). A single-underlier "
                        "fund (one stock or one futures contract, e.g. crypto-linked ETHU/SOLT) shows 1 -- not "
                        "automatically disqualifying if the underlying asset itself is explicitly in scope.",
    "underlier_note": "One-line description of what the fund actually tracks and any notable structural fact "
                       "(ETN issuer credit risk, shared underlier with a sibling ticker, etc.).",
    "candidate_type": "Which candidate-selection method produced this row: 'best safe node' (robust_alpha, "
                       "required to pass cliff-safety -- this project's real selection convention), 'best "
                       "CAGR-safe node' (same cliff-safety check, ranked on alpha_vs_spy/CAGR-equivalent "
                       "instead of robust_alpha), 'best unsafe node' (top robust_alpha regardless of "
                       "cliff-safety), '5min best possible' (top raw 'possible'-resolution alpha, ignoring "
                       "robustness), or 'best certain' (top alpha_vs_spy_certain, the no-guessing fill "
                       "resolution -- can differ from the others when 'possible's optimistic bounce-fill "
                       "assumption picks a different node).",
    "also_matches": "Every candidate_type label that landed on this SAME node (params identical) -- shown as "
                    "'A = B' when two+ methods agree, so a row isn't presented as one method's pick when it's "
                    "really a consensus across several.",
    "node_id": "Stable real id (candidate_nodes.id) for this exact node's full param tuple -- give this ID "
               "back to set pick/comment (survives across report re-runs, unlike a row's position in the "
               "sheet). Blank only if this row's node has no cached backtest data.",
    "pick": "Your decision: yes / no / blank (not yet decided). Persisted in the DB against node_id, so it "
            "carries forward into the next report re-run instead of being wiped every time.",
    "comment": "Your free-text note on this node (e.g. 'SpaceX 2x'). Persisted alongside pick.",
    "strategy": "TrailingBothZScoreBreakout (trailing-buy entry, the live default) or TrailingExitZScoreBreakout "
                "(market-buy entry).",
    "liquidity_dollars_per_day": "avg_vol_10d * last_price * 0.01 -- this project's standard dollar-liquidity "
                                  "estimate. Supposed to be the FIRST-pass filter before spending validation "
                                  "effort on a candidate -- a great alpha number on an illiquid name is "
                                  "untradeable regardless of the rest of this row.",
    "liquidity_tranche": "Liquidity bucketed at $50k/$100k/$500k/$1M boundaries.",
    "core_alpha_pct": "robust_alpha for this row's node: MIN(possible, pessimistic, certain) fill-resolution "
                       "alpha vs SPY, over the ticker's full cached-data window. WITH SPY subtracted.",
    "abs_return_pct": "This node's own RAW compounded return (SPY NOT subtracted) over the same window.",
    "strategy_cagr_pct": "Annualized (CAGR) version of abs_return_pct, over the real cached-data years span.",
    "cagr_tranche": "strategy_cagr_pct bucketed (negative / 0-50% / 50-100% / 100-200% / 200%+).",
    "ann_excess_pct": "CAGR-based excess return over SPY (SPY-adjusted, like core_alpha_pct), annualized -- "
                       "fixes the cross-ticker horizon-mismatch problem (tickers have very different real "
                       "cached-data spans). Always read next to years/trades -- a large value backed by a "
                       "short span/few trades is an annualization artifact, not a real signal.",
    "alpha_possible_pct": "Raw alpha_vs_spy under the 'possible' fill resolution (the kernel's default, "
                           "optimistic-but-unmodified bounce-fill assumption). Empirically the most accurate "
                           "single resolution (docs/research_log.md's 2026-08-08 entries).",
    "alpha_pessimistic_pct": "Alpha under the 'pessimistic' fill resolution (mirror-image Low-after-High "
                              "assumption).",
    "alpha_certain_pct": "Alpha under the 'certain' fill resolution (only resolves a fill when provable "
                          "regardless of intrabar ordering -- no guessing).",
    "resolution_spread_tranche": "How far apart alpha_possible_pct and alpha_certain_pct are (TIGHT/MODERATE/"
                                  "WIDE) -- a wide spread means this node's real number is genuinely uncertain "
                                  "depending on intrabar fill assumptions, not just a rounding difference.",
    "years": "Real calendar span of this ticker's cached hourly data (years). Varies a lot by ticker (e.g. "
             "SOXL ~3.0y vs SPCL ~0.31y) -- core_alpha_pct/abs_return_pct alone aren't fairly comparable "
             "across tickers without checking this.",
    "years_tranche": "years bucketed (<1y / 1-2y / 2-3y / 3y+).",
    "trades": "Number of completed trades in this row's node's backtest.",
    "trades_tranche": "trades bucketed (<10 / 10-25 / 25-50 / 50-100 / 100+). A low count here is the first "
                       "thing to check before trusting a big alpha number -- see core_fluke_verdict too.",
    "worst_neighbor_pct": "Cliff-safety check: the worst robust_alpha found among nearby grid values "
                           "(CLIFF_RADIUS=3 steps on take_profit/stop_loss, +/-7 on max_hold_hours), holding "
                           "every other axis fixed. Negative means a nearby parameter nudge would have lost "
                           "money.",
    "status": "CLIFF (worst_neighbor_pct < 0, fragile to small parameter changes) or SAFE (>= 0), computed "
              "fresh for THIS row's own node.",
    "cliff_tranche": "Clean alias of status, for consistent _tranche column naming in a pivot table.",
    "fillacc_possible_win_pct": "Of this node's real trailing-buy entry signals in the last ~58 days "
                                 "(yfinance's 5-min history cap), the % where the 'possible' fill resolution "
                                 "was closest to the REAL 5-minute-bar fill price. Blank for "
                                 "TrailingExitZScoreBreakout (market-buy has no bounce-fill resolution to "
                                 "check).",
    "fillacc_possible_mean_err_pct": "Mean absolute price error (%) of the 'possible' resolution vs the real "
                                      "5-min fill, across those same signals. Lower is better.",
    "fillacc_n": "How many real signals the fill-accuracy check is based on -- often single digits, treat a "
                 "100% win rate on n=1-2 with real caution.",
    "addon_n": "Number of backtested add-on-leg trades for THIS exact node (a generic, unvalidated confirm_days "
               "config from run_overlay_shim.py -- add-on has no independent parameter to tune, per "
               "docs/overlay_parameter_robustness_process.md's taxonomy).",
    "addon_compounded_pct": "Compounded return of just the add-on overlay's own trades (not combined with "
                             "core).",
    "addon_win_rate_pct": "% of add-on trades that were profitable.",
    "addon_robustness_verdict": "OK / FRAGILE (half negative) / FRAGILE (sign flip) -- does add-on's result "
                                 "hold up across a chronological half1/half2 split and survive removing its "
                                 "single biggest winning trade (overlay_parameter_robustness_process.md steps "
                                 "1 and 3).",
    "addon_tranche": "OK / FRAGILE / NONE bucket of addon_robustness_verdict.",
    "addon_early_wr_pct": "Add-on win rate in the early (first ~70%) half of its chronological trade list.",
    "addon_late_wr_pct": "Add-on win rate in the late (last ~30%) half.",
    "addon_wr_verdict": "STABLE / FADING (>20pp win-rate drop late vs early) / blank (too few trades in a "
                         "half) -- checklist item 4 (win-rate stability), applied to add-on.",
    "addon_wr_tranche": "Clean bucket of addon_wr_verdict (NONE if no overlay data).",
    "drought_n": "Number of backtested drought-overlay trades for THIS exact node (generic confirm_days=10 "
                 "config, NOT the vol-gated/tuned config -- see drought_ie_* for the vol-gate check).",
    "drought_compounded_pct": "Compounded return of just the drought overlay's own trades (not combined with "
                               "core).",
    "drought_win_rate_pct": "% of drought trades that were profitable.",
    "drought_robustness_verdict": "Same OK/FRAGILE check as addon_robustness_verdict, applied to drought.",
    "drought_tranche": "OK / FRAGILE / NONE bucket of drought_robustness_verdict.",
    "drought_early_wr_pct": "Drought win rate in the early (first ~70%) half of its chronological trade list.",
    "drought_late_wr_pct": "Drought win rate in the late (last ~30%) half.",
    "drought_wr_verdict": "STABLE / FADING / blank -- checklist item 4, applied to drought.",
    "drought_wr_tranche": "Clean bucket of drought_wr_verdict.",
    "core_addon_cagr_pct": "Annualized CAGR of core stacked with add-on -- add-on's own factor is used only "
                            "if addon_robustness_verdict is OK, otherwise treated as a no-op (1x, i.e. equal "
                            "to strategy_cagr_pct) so a FRAGILE overlay can't inflate the combined number.",
    "core_drought_cagr_pct": "Same as core_addon_cagr_pct, for drought (gated on drought_robustness_verdict) -- "
                              "EXCEPT when drought_ie_verdict=REAL_SELECTION, in which case the vol-gated "
                              "included-only return replaces the plain drought factor entirely (a confirmed "
                              "real filter is a strictly better estimate than the ungated overlay).",
    "core_both_cagr_pct": "Core stacked with BOTH add-on and drought, each independently gated on its own "
                           "robustness verdict (or the vol-gated REAL_SELECTION override for drought, see "
                           "core_drought_cagr_pct) -- multiplicative, not additive (two ~1.5x-robust overlays "
                           "combine to ~2.3x, not ~3x -- see docs/research_log.md's 2026-08-09 entry).",
    "core_sdb_cagr_pct": "Annualized CAGR of the core node's OWN trade sequence with same_day_block applied -- "
                          "unlike core_addon_cagr_pct/core_drought_cagr_pct (which STACK an overlay on top of "
                          "core), same_day_block instead REPLACES some of core's own trades (blocks a same-day "
                          "cash-account re-buy), so this is CAGR of sdb_compounded_blocked_pct directly, not a "
                          "multiplicative combination with strategy_cagr_pct. Always computed automatically for "
                          "every TrailingBoth candidate (unless --skip-walkforward), matching the drought/addon "
                          "'compute it for every row, not just on request' convention. Blank for "
                          "TrailingExitZScoreBreakout (no same_day_block parameter to test).",
    "drought_ie_confirm_days": "confirm_days used for the included-vs-excluded challenge below (pulled from "
                                "this node's own existing drought overlay run, so it stays consistent with "
                                "drought_n above).",
    "drought_ie_vol_gate": "Entry-time intraday-vol percentile gate tested (default 0.4, the one value "
                            "actually validated live, from SOXL's real confirm_days=3/vol_gate=0.4 signal).",
    "drought_ie_n_included": "How many drought windows the vol gate KEPT (entry-time vol percentile < gate).",
    "drought_ie_included_compounded_pct": "Compounded return of the kept (included) windows only.",
    "drought_ie_included_win_rate_pct": "Win rate of the kept (included) windows only.",
    "drought_ie_n_excluded": "How many drought windows the vol gate THREW OUT (entry-time vol percentile >= "
                              "gate).",
    "drought_ie_excluded_compounded_pct": "Compounded return of the excluded windows -- compare against "
                                           "included to judge whether the filter is doing real work.",
    "drought_ie_excluded_win_rate_pct": "Win rate of the excluded windows.",
    "drought_ie_verdict": "REAL_SELECTION (included beats excluded on both return and win rate, AND included "
                           "is itself profitable) / DISCRIMINATES_BUT_UNPROFITABLE (included beats excluded "
                           "but is still a loser) / NO_REAL_SELECTION / N/A (all one side). "
                           "overlay_parameter_robustness_process.md step 4, drought-only -- add-on has no "
                           "independent filter to challenge.",
    "drought_ie_tranche": "Clean bucket of drought_ie_verdict.",
    "core_fluke_trades": "Number of trades in the core node's own closed-trade list (same as trades column).",
    "core_fluke_alpha_pct": "Core node's alpha vs SPY (see core_alpha_pct) -- repeated here so it sits next "
                             "to the without-biggest-trade version for a direct before/after comparison.",
    "core_fluke_alpha_without_biggest_pct": "Core node's alpha vs SPY with its SINGLE biggest-return trade "
                                             "removed. A large drop, or a sign flip, means the alpha is "
                                             "single-trade-dependent, not a real repeatable edge.",
    "core_fluke_compounded_pct": "Core node's raw compounded return (not alpha) -- for reference alongside "
                                  "the alpha-based fields above.",
    "core_fluke_compounded_without_biggest_pct": "Core node's raw compounded return with the biggest trade "
                                                  "removed.",
    "core_fluke_verdict": "OK / FLUKE (single-trade dependent) -- checklist item 8, tests ALPHA sign (not raw "
                           "return sign, since this whole report ranks on alpha vs SPY).",
    "core_fluke_tranche": "Clean bucket of core_fluke_verdict.",
    "core_wr_early_pct": "Core node's win rate in the early (first ~70%) half of its chronological trade list.",
    "core_wr_late_pct": "Core node's win rate in the late (last ~30%) half.",
    "core_wr_diff_pct": "late minus early win rate -- a large negative number means the edge may be fading.",
    "core_wr_verdict": "STABLE / FADING (>20pp drop) / blank if the late-window sample is under 5 trades (too "
                        "thin to trust a verdict either way) -- checklist item 4, applied to core.",
    "core_wr_tranche": "STABLE / FADING / THIN_SAMPLE / NONE bucket of core_wr_verdict.",
    "wf_verdict": "PASS (all N-fold walk-forward out-of-time folds positive) / MARGINAL (at least half) / FAIL "
                  "/ THIN (fewer folds had trades than requested) -- checklist item 13. Distinct from "
                  "core_wr_verdict: this tests ALPHA consistency across independent time folds, not a single "
                  "win-rate split.",
    "wf_tranche": "PASS / MARGINAL / FAIL / THIN / NONE bucket of wf_verdict.",
    "wf_positive_folds": "How many of the N walk-forward folds had positive alpha.",
    "wf_total_folds": "N (default 5).",
    "wf_min_fold_alpha": "Worst single fold's alpha vs SPY.",
    "wf_max_fold_alpha": "Best single fold's alpha vs SPY.",
    "wf_mean_fold_alpha": "Average fold alpha vs SPY.",
    "max_drawdown_pct": "True peak-to-trough max drawdown across the core node's full compounded equity curve "
                         "-- checklist item 11.",
    "current_drawdown_pct": "Where the equity curve sits RIGHT NOW relative to that worst-case peak -- "
                             "checklist item 12. 0% means currently at a new equity high.",
    "trend_30d_pct": "Real 30-day price return of the underlying ticker, independent of the backtest's "
                      "mean-reversion assumption -- checklist item 1. A large sustained move means recent "
                      "signals are fighting a real trend, not just chopping around a stable mean.",
    "trend_90d_pct": "Real 90-day price return, same idea as trend_30d_pct.",
    "split_flag": "Real yfinance-confirmed stock splits landing inside the ticker's cached date range -- "
                   "checklist item 6. A missed split silently corrupts the cached price series.",
    "sdb_trade_retention_pct": "% of core trades that SURVIVE the real cash-account same-day-re-buy rule "
                                "(schwab_safety.py's same_day_block, enforced live) -- checklist item 9.",
    "sdb_alpha_retention_pct": "% of core ROBUST ALPHA that survives the same same-day-block rule -- can "
                                "diverge sharply from trade retention if the blocked trades happen to be the "
                                "strongest ones. Only meaningful when the unblocked baseline alpha is positive "
                                "(blank otherwise -- see sdb_alpha_unblocked_pct/sdb_alpha_blocked_pct for the "
                                "raw before/after values regardless).",
    "sdb_alpha_unblocked_pct": "Robust alpha with same_day_block OFF (the raw baseline).",
    "sdb_alpha_blocked_pct": "Robust alpha with same_day_block ON.",
    "sdb_compounded_unblocked_pct": "Core node's raw compounded return with same_day_block OFF -- the same "
                                     "kind of number as drought_compounded_pct/addon_compounded_pct, feeding "
                                     "core_sdb_cagr_pct below.",
    "sdb_compounded_blocked_pct": "Core node's raw compounded return with same_day_block ON.",
    "sdb_retention_early_pct": "sdb_alpha_retention_pct computed on just the early half of the same-day-block "
                                "stability split -- checklist item 10.",
    "sdb_retention_late_pct": "sdb_alpha_retention_pct computed on just the late half.",
    "drawdown_tranche": "max_drawdown_pct bucketed (<10% / 10-25% / 25-50% / 50%+).",
    "trend_tranche": "trend_30d_pct bucketed (STRONG_UP / MILD_UP / FLAT / MILD_DOWN / STRONG_DOWN).",
    "split_tranche": "NONE (no splits found) / FLAGGED (a real split lands in the cached window) bucket of "
                      "split_flag.",
    "sdb_tranche": "LOW_RETENTION / MODERATE_RETENTION / HIGH_RETENTION bucket of sdb_alpha_retention_pct. "
                   "'N/A (TrailingExit)' for TrailingExitZScoreBreakout rows -- same_day_block only applies to "
                   "TrailingBothZScoreBreakout.",
    "sdb_recommend": "Whether force_same_day_block looks worth flagging on this node's real watch_list row, "
                     "same framing as the drought/addon overlay verdicts -- a direct 'is this a real execution "
                     "lever worth using' recommendation, not just the raw retention number. BLOCK_HELPS: "
                     "blocked robust alpha exceeds unblocked (retention >105%) AND holds up in both the early "
                     "and late stability halves (checklist item 10) -- the LABU shape (76.6% vs 73.0%). "
                     "BLOCK_HELPS_UNSTABLE: retention >105% overall but doesn't hold in one stability half -- "
                     "real gain, but don't trust it blindly, check the split. NEUTRAL: retention 95-105%, "
                     "not worth the added complexity either way. COSTS: retention <95%, blocking this node "
                     "would give up real alpha. N/A (TrailingExit) / no data otherwise.",
    "x_addon_pct": "NAIVE estimate of core+add-on combined: (1+core_return)*(1+addon_return)-1. Add-on capital "
                   "runs CONCURRENTLY with an open core position (not sequentially), so this OVERSTATES the "
                   "real combined effect -- treat as a rough upper bound, not a real number.",
    "x_drought_pct": "NAIVE estimate of core+drought combined, same formula. Drought fills core's own idle-time "
                      "gaps SEQUENTIALLY, so this approximation is more defensible than x_addon_pct, but still "
                      "not the rigorous stacked model.",
    "exit_fillacc_win_pct": "Exit-side mirror of fillacc_possible_win_pct -- checklist item 3, trailing-SELL "
                             "5-min resolution accuracy. TrailingBothZScoreBreakout only.",
    "exit_fillacc_mean_err_pct": "Exit-side mirror of fillacc_possible_mean_err_pct.",
    "exit_fillacc_n": "How many real armed-trailing-sell signals the exit-side fill-accuracy check is based on.",
    "bear_proxy": "The long-history 1x underlying ETF this ticker's synthetic leveraged price series was "
                  "reconstructed from (e.g. SOXL -> SOXX). Blank if no proxy mapping exists for this ticker.",
    "bear_leverage": "The leverage multiplier applied to the proxy to synthesize this ticker's price (negative "
                      "= inverse product).",
    "bear_note": "Any caveat about this ticker's specific proxy mapping (e.g. a rough high-beta approximation).",
    "bear_worst_crash": "Which of the 4 historical crash windows produced this node's worst COMBINED return "
                         "(0-trade crashes excluded from this comparison -- see bear_market_tranche).",
    "bear_worst_compounded_pct": "The combined-window compounded return in that worst crash.",
    "bear_market_tranche": "POSITIVE / LOSS_<10-50%|50-90%|90%+> bucket of bear_worst_compounded_pct / "
                            "NO_TRADES_ANY_CRASH (every crash window had zero real signals -- distinct from "
                            "surviving unscathed) / NO_PROXY (no long-history underlying exists for this "
                            "ticker) / NONE.",
    "bear_2008_gfc_decline_pct": "SYNTHETIC daily-bar approximation (not the real hourly kernel -- see "
                                  "crash25_* for a real-hourly-kernel check) of this node's compounded return "
                                  "during the 2008 GFC decline leg only (peak to bottom).",
    "bear_2008_gfc_combined_pct": "Same synthetic sim, decline+recovery combined (peak to full SPY recovery).",
    "bear_2008_gfc_max_dd_pct": "Peak-to-trough drawdown of this node's OWN synthetic equity curve during the "
                                 "combined window -- the single most decision-relevant number here.",
    "bear_2008_gfc_trades": "How many synthetic trades fired during the combined window -- 0 means this node "
                             "never triggered during this crash, not that it broke even (see "
                             "bear_market_tranche's NO_TRADES_ANY_CRASH note).",
    "bear_2020_covid_decline_pct": "Same as bear_2008_gfc_decline_pct, for the 2020 COVID crash.",
    "bear_2020_covid_combined_pct": "Same as bear_2008_gfc_combined_pct, for 2020 COVID.",
    "bear_2020_covid_max_dd_pct": "Same as bear_2008_gfc_max_dd_pct, for 2020 COVID.",
    "bear_2020_covid_trades": "Same as bear_2008_gfc_trades, for 2020 COVID.",
    "bear_2022_bear_decline_pct": "Same as bear_2008_gfc_decline_pct, for the 2022 bear market.",
    "bear_2022_bear_combined_pct": "Same as bear_2008_gfc_combined_pct, for 2022 bear.",
    "bear_2022_bear_max_dd_pct": "Same as bear_2008_gfc_max_dd_pct, for 2022 bear.",
    "bear_2022_bear_trades": "Same as bear_2008_gfc_trades, for 2022 bear.",
    "bear_2000_dotcom_decline_pct": "Same as bear_2008_gfc_decline_pct, for the 2000 dotcom crash -- CHECK "
                                     "bear_2000_dotcom_truncated first, several proxies postdate this crash's "
                                     "real decline leg.",
    "bear_2000_dotcom_combined_pct": "Same as bear_2008_gfc_combined_pct, for 2000 dotcom.",
    "bear_2000_dotcom_max_dd_pct": "Same as bear_2008_gfc_max_dd_pct, for 2000 dotcom.",
    "bear_2000_dotcom_trades": "Same as bear_2008_gfc_trades, for 2000 dotcom.",
    "bear_2000_dotcom_truncated": "True if this ticker's underlying proxy's real data history starts AFTER "
                                   "the official 2000 dotcom decline leg began (e.g. SOXX launched 2001-07, "
                                   "well after the 2000-03 peak) -- when True, the dotcom columns above cover "
                                   "a truncated/later window, not the real full crash.",
    "crash25_ticker_bh_decline_pct": "REAL (not synthetic) buy-and-hold return of the ticker itself, off real "
                                      "intraday hourly bars, from the real 2025 SPY peak (2025-02-19 15:30) to "
                                      "the real SPY trough (2025-04-07 10:30).",
    "crash25_ticker_bh_recovery_pct": "Real buy-and-hold return from the SPY trough to the day SPY's own price "
                                       "first reclaimed its pre-crash high (2025-06-27 09:30).",
    "crash25_ticker_bh_combined_pct": "Real buy-and-hold return across the full window (peak to SPY recovery).",
    "crash25_ticker_bh_max_dd_pct": "Real intrabar (hourly-bar, not daily-close) max drawdown of simply "
                                     "holding the ticker across the full window.",
    "crash25_algo_decline_pct": "This node's REAL hourly-kernel trade record (the actual backtest, not a "
                                 "synthetic approximation) compounded return during the decline leg only.",
    "crash25_algo_recovery_pct": "Same, for the recovery leg (SPY trough to SPY recovery).",
    "crash25_algo_combined_pct": "Same, across the full window -- compare directly against "
                                  "crash25_ticker_bh_combined_pct to see if the algo beat simply holding "
                                  "through this real crash.",
    "crash25_algo_max_dd_pct": "This node's OWN real max drawdown across the full window -- compare against "
                                "crash25_ticker_bh_max_dd_pct.",
    "crash25_algo_trades": "How many real trades this node's kernel produced during the window.",
    "crash25_verdict": "ALGO_OUTPERFORMED_BH / ALGO_UNDERPERFORMED_BH -- did the algo's real combined return "
                        "beat simply holding the ticker through the real 2025 crash+recovery?",
    "crash_2025_tranche": "Clean bucket of crash25_verdict. NONE if this ticker's cached history doesn't reach "
                           "back to the 2025-02-19 crash peak (real gap for several tickers, e.g. GDXU/USD/"
                           "NUGT/BOIL/JNUG/UGL all start exactly 2025-06-27).",
}

FIELDNAMES = list(COLUMN_DEFS.keys())


def exit_fill_accuracy_summary(ticker, strategy, node):
    """Checklist item 3: trailing-SELL resolution check, the exit-side mirror
    of item 2's entry-side fillacc columns (already reused as-is from
    candidate_summary_report.py). TrailingBothZScoreBreakout only -- exit
    logic for TrailingExitZScoreBreakout is a plain market-sell, no trailing-
    stop resolution to check. Real yfinance 5-min replay, same
    FIVE_MIN_LOOKBACK_DAYS window as the entry-side check, so this only ever
    covers recently-armed trades -- can legitimately be (win%, err%, n=0) for
    a node that hasn't armed in the last ~58 days."""
    if strategy != "TrailingBothZScoreBreakout":
        return None
    cutoff = _dt.now() - _timedelta(days=FIVE_MIN_LOOKBACK_DAYS)
    try:
        events = find_hourly_trailing_exits(
            ticker, node["window"], node["z"], node["trail_buy_pct"] / 100.0,
            node["trail_sell_pct"] / 100.0, node["arm_pct"] / 100.0, node["hold"], cutoff)
    except Exception:
        return None
    if not events:
        return (None, None, 0)
    try:
        df_5m = yf.Ticker(ticker).history(period=f"{FIVE_MIN_LOOKBACK_DAYS}d", interval="5m")
        if df_5m.index.tz is not None:
            df_5m.index = df_5m.index.tz_localize(None)
    except Exception:
        return None
    diffs = []
    for ev in events:
        real = replay_five_min(df_5m, ev["arm_time"], ev["peak_at_arm"], node["trail_sell_pct"] / 100.0,
                                ev["cutoff_time"])
        if real is None:
            continue
        diffs.append((real["five_min_exit_price"] - ev["hourly_exit_price"]) / ev["hourly_exit_price"] * 100)
    if not diffs:
        return (None, None, 0)
    mean_abs_err = sum(abs(d) for d in diffs) / len(diffs)
    win_rate = sum(1 for d in diffs if abs(d) < 0.5) / len(diffs) * 100  # within noise, same convention as item 2
    return (win_rate, mean_abs_err, len(diffs))

_TREND_CACHE = {}
_SPLIT_CACHE = {}

# Real (not synthetic) SPY drawdown -- confirmed directly against cached SPY
# HOURLY bars (intraday, not daily-resampled -- a daily Close-to-Close
# series understates both the true peak/trough and the true magnitude,
# since it misses whatever happened intrabar): peak 2025-02-19 15:30
# ($612.90), trough 2025-04-07 10:30 ($493.79, -19.4%), SPY's hourly Close
# first reclaimed the peak level on 2025-06-27 09:30. Distinct from
# bear_market_stress_check() above, which uses a SYNTHETIC daily-bar
# approximation of the strategy against decades-old proxy data -- this
# check runs the REAL hourly-bar kernel (the same trade list
# core_walk_forward() already built) against a real, in-cache crash the
# live strategy could actually have traded through, and benchmarks against
# each ticker's own real intraday buy-hold price action over the same
# window, not a daily-close approximation of it. A shared calendar anchor
# (SPY's own recovery timestamp, not each ticker's OWN price recovery date)
# is used across every ticker so results are comparable to each other, not
# just to each ticker's own history.
CRASH_2025_PEAK = pd.Timestamp("2025-02-19 15:30:00")
CRASH_2025_TROUGH = pd.Timestamp("2025-04-07 10:30:00")
CRASH_2025_SPY_RECOVERY = pd.Timestamp("2025-06-27 09:30:00")

_HOURLY_PRICE_CACHE = {}


def _hourly_close_series(ticker):
    if ticker in _HOURLY_PRICE_CACHE:
        return _HOURLY_PRICE_CACHE[ticker]
    path = CACHE_DIR / f"{ticker}_1h.csv"
    if not path.exists():
        _HOURLY_PRICE_CACHE[ticker] = None
        return None
    df = pd.read_csv(path, index_col=0, parse_dates=True).sort_index()
    close_col = "Adj Close" if "Adj Close" in df.columns else "Close"
    _HOURLY_PRICE_CACHE[ticker] = df[close_col]
    return df[close_col]


def _price_at_or_before(series, ts):
    prior = series[series.index <= ts]
    return float(prior.iloc[-1]) if len(prior) else None


def real_crash_2025_check(ticker, closed_trades):
    """Real (not synthetic) 2025 drawdown/recovery check -- see
    CRASH_2025_PEAK/_TROUGH/_SPY_RECOVERY above. `closed_trades` is the SAME
    chronological closed-trade list core_walk_forward() already built via
    the real hourly kernel (t_base), so this is a windowed re-slice of
    already-computed data, not a second kernel run.

    Returns None if this ticker's own cached history doesn't reach back to
    CRASH_2025_PEAK (real gap -- GDXU/USD/NUGT/BOIL/JNUG/UGL among others
    have zero real data before 2025-06-27, so this check is honestly
    inapplicable for them, not zero-filled). Else a dict:
      {ticker_bh_decline_pct, ticker_bh_recovery_pct, ticker_bh_combined_pct,
       ticker_bh_max_dd_pct, algo_decline_pct, algo_recovery_pct,
       algo_combined_pct, algo_max_dd_pct, algo_trades, verdict}
    verdict compares algo_combined_pct against ticker_bh_combined_pct (has
    the algo's own real trade record over this specific real crash beaten
    simply holding the ticker through it, as of the day SPY got back to even).
    """
    hourly = _hourly_close_series(ticker)
    if hourly is None or hourly.index.min() > CRASH_2025_PEAK:
        return None

    p_peak = _price_at_or_before(hourly, CRASH_2025_PEAK)
    p_trough = _price_at_or_before(hourly, CRASH_2025_TROUGH)
    p_recovery = _price_at_or_before(hourly, CRASH_2025_SPY_RECOVERY)
    if p_peak is None or p_trough is None or p_recovery is None:
        return None

    # Real intrabar drawdown, off the same hourly bars the live strategy
    # actually sees -- NOT resampled to one Close-per-day, which would
    # understate both how deep and how fast this ticker's own real crash was.
    window_series = hourly[(hourly.index >= CRASH_2025_PEAK) & (hourly.index <= CRASH_2025_SPY_RECOVERY)]
    bh_running_max = window_series.cummax()
    bh_dd = ((window_series - bh_running_max) / bh_running_max * 100)
    ticker_bh_max_dd_pct = float(bh_dd.min()) if len(bh_dd) else None

    ticker_bh_decline_pct = (p_trough / p_peak - 1) * 100
    ticker_bh_recovery_pct = (p_recovery / p_trough - 1) * 100
    ticker_bh_combined_pct = (p_recovery / p_peak - 1) * 100

    window_trades = [t for t in closed_trades
                      if CRASH_2025_PEAK <= t["Entry Time"] <= CRASH_2025_SPY_RECOVERY]
    decline_trades = [t for t in window_trades if t["Entry Time"] < CRASH_2025_TROUGH]
    recovery_trades = [t for t in window_trades if t["Entry Time"] >= CRASH_2025_TROUGH]

    algo_decline_pct = compounded([t["Return"] for t in decline_trades]) if decline_trades else 0.0
    algo_recovery_pct = compounded([t["Return"] for t in recovery_trades]) if recovery_trades else 0.0

    equity, peak_equity, max_dd = 1.0, 1.0, 0.0
    for t in window_trades:
        equity *= (1.0 + t["Return"])
        peak_equity = max(peak_equity, equity)
        max_dd = min(max_dd, (equity - peak_equity) / peak_equity)
    algo_combined_pct = (equity - 1.0) * 100
    algo_max_dd_pct = max_dd * 100

    verdict = ("ALGO_OUTPERFORMED_BH" if algo_combined_pct > ticker_bh_combined_pct
               else "ALGO_UNDERPERFORMED_BH")

    return {
        "ticker_bh_decline_pct": ticker_bh_decline_pct, "ticker_bh_recovery_pct": ticker_bh_recovery_pct,
        "ticker_bh_combined_pct": ticker_bh_combined_pct, "ticker_bh_max_dd_pct": ticker_bh_max_dd_pct,
        "algo_decline_pct": algo_decline_pct, "algo_recovery_pct": algo_recovery_pct,
        "algo_combined_pct": algo_combined_pct, "algo_max_dd_pct": algo_max_dd_pct,
        "algo_trades": len(window_trades), "verdict": verdict,
    }


def ticker_trend(ticker):
    """Checklist item 1: 30d/90d return, independent of the backtest's mean-
    reversion assumption -- a large sustained move means recent signals are
    fighting a real trend, not just chopping around a stable mean. Cached
    per-ticker (cheap CSV read, but no need to redo it per candidate row)."""
    if ticker in _TREND_CACHE:
        return _TREND_CACHE[ticker]
    path = CACHE_DIR / f"{ticker}_1h.csv"
    if not path.exists():
        _TREND_CACHE[ticker] = (None, None)
        return None, None
    df = pd.read_csv(path, index_col=0, parse_dates=True)
    close_col = "Adj Close" if "Adj Close" in df.columns else "Close"
    daily = df.resample("D").last().dropna(subset=[close_col])
    if len(daily) < 64:
        result = (None, None)
    else:
        r30 = (daily[close_col].iloc[-1] / daily[close_col].iloc[-21] - 1) * 100
        r90 = (daily[close_col].iloc[-1] / daily[close_col].iloc[-63] - 1) * 100
        result = (float(r30), float(r90))
    _TREND_CACHE[ticker] = result
    return result


def ticker_split_flag(ticker):
    """Checklist item 6: real yfinance-confirmed stock splits landing inside
    the ticker's cached date range -- a missed split silently corrupts the
    cached price series. Cached per-ticker (real network call)."""
    if ticker in _SPLIT_CACHE:
        return _SPLIT_CACHE[ticker]
    path = CACHE_DIR / f"{ticker}_1h.csv"
    if not path.exists():
        _SPLIT_CACHE[ticker] = None
        return None
    try:
        splits = _check_splits(path)
        result = "; ".join(f"{s['split_date']} ({s['ratio']}x)" for s in splits) if splits else "none"
    except Exception as e:
        result = f"check failed: {e}"
    _SPLIT_CACHE[ticker] = result
    return result


def candidate_type_membership(df_t, min_alpha):
    """Independently computes the four raw candidate-selection node keys
    (not deduped, unlike find_candidates()) so a collapsed row can be
    labeled with EVERY type that landed on it, not just whichever label
    find_candidates() happened to keep. Answers "which ones overlapped?"
    for any row with fewer than 4 distinct candidates for its ticker.

    Also returns raw_label -> full node dict (window/z/sl/arm_pct/
    trail_buy_pct/trail_sell_pct/hold/entry_timing, plus the real
    possible/pessimistic/certain alphas) for every candidate computed here --
    the SAME node objects `find_candidates()` itself would produce, so a row
    can look its own params up directly instead of round-tripping through a
    `WHERE robust_alpha=?` float-equality DB query (found 2026-08-09, paired
    review: that query could silently return no match after any
    backtest_cache recompute, or match the WRONG node on a robust_alpha tie
    between two grid plateaus -- this makes that whole lookup unnecessary)."""
    membership = {}  # node_key -> [label, ...]
    label_to_node = {}  # raw_label -> node dict

    cliff_safe = best_safe_node(df_t, min_alpha=min_alpha, metric="robust_alpha")
    if cliff_safe:
        cliff_safe['robust_alpha'] = cliff_safe.pop('alpha')
        membership.setdefault(_node_key(cliff_safe), []).append('best safe node')
        label_to_node['cliff-safe (current convention)'] = cliff_safe

    # CAGR-first pick, added 2026-08-11 per explicit request ("CAGR is more
    # important"/"use that instead of alpha across the board"). alpha_vs_spy
    # (the 'possible' resolution alone) is order-equivalent to CAGR for a
    # fixed ticker -- same years, same SPY baseline apply to every row of its
    # grid -- so reusing best_safe_node() with metric="alpha_vs_spy" ranks
    # AND cliff-checks consistently on the CAGR-equivalent metric, rather
    # than mixing a CAGR-based rank with a robust_alpha-based safety check.
    # Deliberately additive, not a replacement of "best safe node" above --
    # fully swapping robust_alpha for possible-only as the DEFAULT
    # safety/selection metric project-wide is the bigger, already-flagged
    # (2026-08-08/2026-08-10 backlog) convention change that needs the full
    # backtest-change-rollout process, not a quick edit here.
    cagr_safe = best_safe_node(df_t, min_alpha=min_alpha, metric="alpha_vs_spy")
    if cagr_safe:
        # best_safe_node()'s 'alpha' key is always robust_alpha regardless of
        # `metric` (deliberate -- lets a caller see how a metric-X-selected
        # node scores on the always-conservative measure too, same pattern
        # compare_fill_resolution_selection.py already relies on). The actual
        # ranking/safety-check metric used here (alpha_vs_spy, CAGR-order-
        # equivalent) is separately in 'alpha_raw' -- don't relabel 'alpha'
        # as alpha_vs_spy, that would silently swap in the wrong value.
        cagr_safe['robust_alpha'] = cagr_safe.pop('alpha')
        membership.setdefault(_node_key(cagr_safe), []).append('best CAGR-safe node')
        label_to_node['CAGR-first (possible-resolution, cliff-checked on the same metric)'] = cagr_safe

    top_robust_row = df_t.sort_values("robust_alpha", ascending=False).iloc[0]
    top_robust = _node_from_row(top_robust_row)
    membership.setdefault(_node_key(top_robust), []).append('best unsafe node')
    label_to_node['best robust_alpha (ignoring cliff-safety)'] = top_robust

    top_possible_row = df_t.sort_values("alpha_vs_spy", ascending=False).iloc[0]
    top_possible = _node_from_row(top_possible_row)
    membership.setdefault(_node_key(top_possible), []).append('5min best possible')
    label_to_node['best possible (raw alpha_vs_spy)'] = top_possible

    df_cert = df_t.copy()
    df_cert["_alpha_certain_filled"] = df_cert["alpha_vs_spy_certain"].fillna(df_cert["alpha_vs_spy"])
    top_certain_row = df_cert.sort_values("_alpha_certain_filled", ascending=False).iloc[0]
    top_certain = _node_from_row(top_certain_row)
    membership.setdefault(_node_key(top_certain), []).append('best certain')
    label_to_node['best certain (alpha_vs_spy_certain, no-guessing resolution)'] = top_certain

    return membership, label_to_node


def compounded(rets):
    prod = 1.0
    for r in rets:
        prod *= (1 + r)
    return (prod - 1) * 100


def win_rate(rets):
    if not rets:
        return 0.0
    return sum(1 for r in rets if r > 0) / len(rets) * 100


def overlay_robustness(conn, ticker, strategy, version, mechanism, node):
    """Looks up the exact candidate_node_id for this row's params (same match
    as candidate_summary_report.overlay_summary_for_node), pulls its
    chronologically-ordered trades, and runs the mechanical robustness
    checks from docs/overlay_parameter_robustness_process.md that apply to
    an already-generated trade list (steps 1 and 3, PLUS checklist item 4's
    early/late win-rate comparison on the same chronological split -- no
    fit-half parameter search here, since drought/addon use a fixed generic
    config via run_overlay_shim.py, not a per-candidate search).

    Returns None if <2 trades (nothing to split/stress-test), else a dict:
      {half1_pct, half2_pct, flips_sign, verdict,
       early_win_rate_pct, late_win_rate_pct, win_rate_diff_pct, win_rate_verdict}
    verdict is one of: 'OK', 'FRAGILE (half negative)', 'FRAGILE (sign flip)'.
    win_rate_verdict is one of: 'STABLE', 'FADING', None (too few trades).
    """
    c = conn.cursor()
    # Real gap found in paired review 2026-08-09: run_overlay_shim.py's INSERT
    # has no dedup, and ensure_overlay_for_node re-runs a mechanism whenever
    # it looks missing -- 22+ real candidate_node_ids in the live DB have
    # 2-5 duplicate run_timestamps for the same mechanism (e.g. LABD node 29
    # addon: 5 runs x 70 identical trades = 350 rows), which silently
    # inflates every stat computed off an un-filtered SELECT here (and
    # neuters the single-trade-removal check below -- removing one copy of
    # the "biggest" trade leaves N-1 identical copies behind). Scoped to the
    # latest run_timestamp per (candidate_node_id, mechanism) so this always
    # reflects one real run, never a blend of duplicates.
    c.execute("""
        SELECT cor.entry_time, cor.ret
        FROM candidate_overlay_results cor
        JOIN candidate_nodes cn ON cn.id = cor.candidate_node_id
        WHERE cn.ticker=? AND cor.mechanism=? AND cn.strategy=? AND cn.version=?
              AND cn.window=? AND cn.z=? AND cn.fixed_sl=? AND cn.arm_pct=?
              AND cn.trail_buy_pct=? AND cn.trail_sell_pct=? AND cn.max_hold_hours=? AND cn.entry_timing=?
              AND cor.run_timestamp = (
                  SELECT MAX(cor2.run_timestamp) FROM candidate_overlay_results cor2
                  WHERE cor2.candidate_node_id = cor.candidate_node_id AND cor2.mechanism = cor.mechanism
              )
        ORDER BY cor.entry_time
    """, (ticker, mechanism, strategy, version, node['window'], node['z'], node['sl'], node['arm_pct'],
          node['trail_buy_pct'], node['trail_sell_pct'], node['hold'], node['entry_timing']))
    trades = c.fetchall()
    if len(trades) < 2:
        return None

    rets = [t[1] for t in trades]
    mid = len(rets) // 2
    half1, half2 = rets[:mid], rets[mid:]
    half1_pct, half2_pct = compounded(half1), compounded(half2)

    comp_all = compounded(rets)
    biggest_idx = max(range(len(rets)), key=lambda i: rets[i])
    without_biggest = rets[:biggest_idx] + rets[biggest_idx + 1:]
    comp_without = compounded(without_biggest) if without_biggest else 0.0
    flips = (comp_all > 0) != (comp_without > 0)

    if flips:
        verdict = "FRAGILE (sign flip)"
    elif half1_pct < 0 or half2_pct < 0:
        verdict = "FRAGILE (half negative)"
    else:
        verdict = "OK"

    # Checklist item 4 (win-rate stability), same chronological split as
    # steps 1/3 above -- reuses half1/half2 rather than a third split.
    early_wr, late_wr = win_rate(half1), win_rate(half2)
    win_rate_diff = late_wr - early_wr
    win_rate_verdict = None if len(half1) < 2 or len(half2) < 2 else (
        "FADING" if win_rate_diff < -20 else "STABLE")

    return {"half1_pct": half1_pct, "half2_pct": half2_pct, "flips_sign": flips, "verdict": verdict,
            "early_win_rate_pct": early_wr, "late_win_rate_pct": late_wr,
            "win_rate_diff_pct": win_rate_diff, "win_rate_verdict": win_rate_verdict}


_IVOL_CACHE = {}


def _get_ivol_series(ticker):
    if ticker not in _IVOL_CACHE:
        try:
            _IVOL_CACHE[ticker] = get_ivol_series(ticker)
        except Exception as e:
            print(f"_get_ivol_series({ticker}): {e}", file=sys.stderr)
            _IVOL_CACHE[ticker] = None
    return _IVOL_CACHE[ticker]


def drought_included_excluded_check(conn, ticker, strategy, version, node, vol_gate=DEFAULT_VOL_GATE):
    """docs/overlay_parameter_robustness_process.md step 4 -- distinct from
    step 3's single-trade-removal: confirm a filter is doing real
    differential selection by comparing what it KEPT (included) against what
    it THREW OUT (excluded), not just eyeballing the included set alone.

    Drought-only. Addon has no independent filter to challenge -- per the
    same doc's taxonomy, add-on inherits core's own trigger/sizing entirely
    (100%-of-position margin, always), so there's no parameter search or
    filter step for it in the first place.

    Uses the entry-time intraday-vol-percentile gate (drought_overlay_sweep.py's
    real get_ivol_series/_entry_vol_pctile, reused directly -- not
    reimplemented) at vol_gate=DEFAULT_VOL_GATE, the one value actually
    validated live (SOXL's confirm_days=3/vol_gate=0.4 signal,
    docs/research_log.md's 2026-08-07 entry) -- run_overlay_shim.py's own
    drought computation (which feeds drought_n/drought_compounded_pct
    elsewhere in this report) does NOT apply this gate at all, so this is a
    genuinely separate computation, not a re-read of existing rows.

    confirm_days is pulled from this node's own existing candidate_overlay_results
    rows (whatever run_overlay_shim.py was last run with) rather than
    re-guessing a default, so the drought windows being challenged here match
    the ones already reported elsewhere on the same row.

    Returns None if <2 real trades or <2 drought windows exist. Else a dict:
      {confirm_days, vol_gate, n_included, included_compounded_pct, included_win_rate_pct,
       n_excluded, excluded_compounded_pct, excluded_win_rate_pct, verdict}
    verdict is one of: 'REAL_SELECTION' (discriminates AND included is profitable),
    'DISCRIMINATES_BUT_UNPROFITABLE' (included beats excluded but is still a
    loser), 'NO_REAL_SELECTION', 'N/A (all one side)'.
    """
    c = conn.cursor()
    # ORDER BY run_timestamp DESC (not a bare LIMIT 1, found in paired review
    # 2026-08-09) -- picks the SAME latest run overlay_robustness()'s
    # duplicate-dedup above now scopes to, so this stays consistent with
    # drought_n/drought_compounded_pct on the same row. confirm_days is
    # nullable in the schema (addon rows write NULL) -- explicitly guarded
    # below rather than trusting `row[0] if row else 10`, which would pass a
    # real NULL straight into find_drought_windows and crash the whole
    # report (TypeError comparing None) outside this function's try/except.
    c.execute("""
        SELECT confirm_days FROM candidate_overlay_results cor
        JOIN candidate_nodes cn ON cn.id = cor.candidate_node_id
        WHERE cn.ticker=? AND cor.mechanism='drought' AND cn.strategy=? AND cn.version=?
              AND cn.window=? AND cn.z=? AND cn.fixed_sl=? AND cn.arm_pct=?
              AND cn.trail_buy_pct=? AND cn.trail_sell_pct=? AND cn.max_hold_hours=? AND cn.entry_timing=?
        ORDER BY cor.run_timestamp DESC
        LIMIT 1
    """, (ticker, strategy, version, node['window'], node['z'], node['sl'], node['arm_pct'],
          node['trail_buy_pct'], node['trail_sell_pct'], node['hold'], node['entry_timing']))
    row = c.fetchone()
    confirm_days = row[0] if row and row[0] is not None else 10  # run_overlay_shim.py's own CLI default

    # get_trades_and_bars/find_drought_windows expect the locate_best_node.py
    # node_dict() key shape (ticker/strategy/fixed_sl/max_hold_hours), NOT
    # this row's candidate_5min_report._node_from_row shape (sl/hold, no
    # ticker/strategy) -- same distinction run_overlay_shim.py's node already
    # gets right; found by a real crash during testing (KeyError: 'strategy').
    gt_node = {
        "ticker": ticker, "strategy": strategy, "window": node["window"], "z": node["z"],
        "fixed_sl": node["sl"], "arm_pct": node["arm_pct"], "trail_buy_pct": node["trail_buy_pct"],
        "trail_sell_pct": node["trail_sell_pct"], "max_hold_hours": node["hold"],
        "entry_timing": node["entry_timing"],
    }
    try:
        trades, df_h = get_trades_and_bars(gt_node)
    except Exception as e:
        print(f"drought_included_excluded_check({ticker}): get_trades_and_bars failed: {e}", file=sys.stderr)
        return None
    if len(trades) < 2:
        return None
    windows = find_drought_windows(trades, df_h, confirm_days)
    if len(windows) < 2:
        return None
    ivol_series = _get_ivol_series(ticker)
    if ivol_series is None:
        return None

    included, excluded = [], []
    for entry_i, gap_end in windows:
        entry_time = df_h.index[entry_i + 1] if entry_i + 1 < len(df_h) else df_h.index[entry_i]
        pctile = _entry_vol_pctile(entry_time, ivol_series)
        if pctile is None:
            continue
        ret = simulate_overlay(df_h, entry_i, gap_end, node['sl'], node['arm_pct'], node['trail_sell_pct'])["ret"]
        (included if pctile < vol_gate else excluded).append(ret)

    if len(included) < 1 or len(excluded) < 1:
        return {"confirm_days": confirm_days, "vol_gate": vol_gate,
                "n_included": len(included), "n_excluded": len(excluded),
                "verdict": "N/A (all one side)"}

    inc_comp, exc_comp = compounded(included), compounded(excluded)
    inc_wr, exc_wr = win_rate(included), win_rate(excluded)
    # Requires the included (kept) set to actually be PROFITABLE, not just
    # better than excluded (found in paired review 2026-08-09: "included
    # -5% vs excluded -40%" discriminated correctly but is still a losing
    # overlay -- reading REAL_SELECTION as a promotion-worthy pass would
    # have been wrong).
    discriminates = (inc_comp > exc_comp) and (inc_wr > exc_wr)
    if discriminates and inc_comp > 0:
        verdict = "REAL_SELECTION"
    elif discriminates:
        verdict = "DISCRIMINATES_BUT_UNPROFITABLE"
    else:
        verdict = "NO_REAL_SELECTION"

    return {
        "confirm_days": confirm_days, "vol_gate": vol_gate,
        "n_included": len(included), "included_compounded_pct": inc_comp, "included_win_rate_pct": inc_wr,
        "n_excluded": len(excluded), "excluded_compounded_pct": exc_comp, "excluded_win_rate_pct": exc_wr,
        "verdict": verdict,
    }


_INPUTS_CACHE = {}


def core_walk_forward(ticker, strategy, node, n_folds=DEFAULT_FOLDS):
    """N-fold out-of-time walk-forward for THIS row's exact core node, reusing
    walk_forward_check.py's walk_forward()/summarize() directly instead of
    reimplementing the fold-slicing/dispersion math. Returns None if this
    ticker/strategy has no cached hourly data. TrailingExitZScoreBreakout
    nodes only get the 'possible' resolution (see module docstring) -- passed
    to walk_forward() as pessimistic=None/certain=None, matching how it
    already handles a missing bound (only includes non-empty lists in the
    per-fold MIN)."""
    strategy_class = STRATEGY_CLASSES.get(strategy)
    if strategy_class is None:
        return None

    # window matters -- prep embeds the daily SMA/Std arrays built with THIS
    # window, so a cache keyed without it silently reuses one candidate's
    # indicators for another's walk-forward/drawdown/same-day-block checks
    # (found 2026-08-09, paired review: 15+ tickers have candidate rows
    # spanning window=10 and window=20, e.g. GDXD -- confirmed materially
    # different real numbers, not just noise).
    cache_key = (ticker, strategy, node["window"])
    inputs = _INPUTS_CACHE.get(cache_key)
    if inputs is None:
        loaded = _load_node_inputs(ticker, strategy_class, strategy, node["window"], node["z"])
        if loaded is None:
            _INPUTS_CACHE[cache_key] = False
            return None
        _, _, prep = loaded
        hourly_path = CACHE_DIR / f"{ticker}_1h.csv"
        df_t = pd.read_csv(hourly_path, index_col=0, parse_dates=True).sort_index()
        if df_t.index.tz is not None:
            df_t.index = df_t.index.tz_localize(None)
        inputs = (prep, df_t.index.min(), df_t.index.max())
        _INPUTS_CACHE[cache_key] = inputs
    if inputs is False:
        return None
    prep, data_start, data_end = inputs

    return_bounds = strategy == "TrailingBothZScoreBreakout"
    result = run_backtest_dispatch(
        strategy_class, df_hourly=None, df_daily_indicators=None, ticker=ticker,
        take_profit=node["arm_pct"], sl_raw=node["trail_buy_pct"] if return_bounds else node["trail_sell_pct"],
        max_hours_to_hold=node["hold"], z_score_threshold=node["z"],
        fixed_sl=node["sl"], trail_pct_pct=node["trail_sell_pct"],
        entry_timing=node["entry_timing"], return_bounds=return_bounds, prep=prep,
    )
    if return_bounds:
        t_base, p_base, c_base = result
    else:
        t_base, p_base, c_base = result, None, None

    fold_rows = walk_forward(t_base, p_base, c_base, data_start, data_end, n_folds)
    if fold_rows is None:
        return None
    summary = summarize_folds(fold_rows)
    if summary.get("folds_with_trades", 0) == 0:
        return None
    positive_folds = summary["folds_with_trades"] - summary["negative_folds"]
    fwt = summary["folds_with_trades"]
    # Indicative only, not a hard gate -- PASS/MARGINAL/FAIL describe fold
    # consistency the same plain-language way `status` (SAFE/CLIFF) already
    # does for cliff-safety, per the user's explicit "these are indicative,
    # some are harder stops than others" framing (2026-08-09). A ticker with
    # fewer folds actually carrying trades than requested (fwt < n_folds) is
    # flagged THIN regardless of the win/loss split -- too little data in
    # some windows to trust the split either way.
    if fwt < n_folds:
        verdict = "THIN (some folds had no trades)"
    elif positive_folds == fwt:
        verdict = "PASS (all folds positive)"
    elif positive_folds >= (fwt / 2):
        verdict = f"MARGINAL ({positive_folds}/{fwt} folds positive)"
    else:
        verdict = f"FAIL ({positive_folds}/{fwt} folds positive)"
    # Checklist items 11/12: true peak-to-trough max drawdown across the
    # node's full compounded equity curve, plus where it sits RIGHT NOW
    # relative to that worst case -- reuses v4_max_drawdown.py's own
    # max_drawdown() against the SAME trade list walk_forward already
    # computed above, no extra kernel run.
    closed = [t for t in t_base if t["Result"] in ("WIN", "LOSS", "TWIN", "TLOSS")]
    max_dd_pct, _, _ = max_drawdown(closed) if closed else (None, None, None)

    crash_2025_check = real_crash_2025_check(ticker, closed)

    # Checklist item 8: is CORE's own "best alpha" driven by a single outlier
    # trade? Mirrors overlay_robustness's single-trade-removal stress test
    # (drought/addon already get this there) -- core never had an equivalent
    # until this pass (found 2026-08-09 follow-up: the trades/trades_tranche
    # count bucket alone doesn't catch a low-count node whose sign flips
    # without its one biggest winner).
    #
    # Tests ALPHA sign (compounded return minus SPY buy-hold over the SAME
    # period), not raw compounded-return sign (found in paired review
    # 2026-08-09) -- this whole report ranks nodes on alpha vs SPY, and SPY
    # was strongly positive over the ~3y cached window, so a node whose
    # compounded return stays positive without its biggest trade but whose
    # ALPHA has actually gone negative would otherwise read as a false "OK".
    fluke_check = None
    if len(closed) >= 2:
        spy_bh_fluke = period_spy_bh(data_start, data_end)
        core_biggest_idx = max(range(len(closed)), key=lambda i: closed[i]["Return"])
        core_without_biggest = closed[:core_biggest_idx] + closed[core_biggest_idx + 1:]
        alpha_all = _summarize_trades(closed, spy_bh_fluke)[0]
        alpha_without = _summarize_trades(core_without_biggest, spy_bh_fluke)[0] if core_without_biggest else -spy_bh_fluke
        core_flips = (alpha_all > 0) != (alpha_without > 0)
        fluke_check = {
            "trades": len(closed), "alpha_pct": alpha_all, "alpha_without_biggest_pct": alpha_without,
            "compounded_pct": compounded([t["Return"] for t in closed]),
            "compounded_without_biggest_pct": (compounded([t["Return"] for t in core_without_biggest])
                                                if core_without_biggest else 0.0),
            "flips_sign": core_flips,
            "verdict": "FLUKE (single-trade dependent)" if core_flips else "OK",
        }

    # Checklist item 4: CORE win-rate stability (early 70% vs late 30%,
    # chronological) -- drought/addon already get the equivalent via
    # overlay_robustness's early_win_rate_pct/late_win_rate_pct (found
    # 2026-08-09 follow-up: this report originally only applied item 4 to
    # core via the fold-alpha walk-forward above, which answers a related
    # but different question -- alpha consistency, not this specific
    # win-RATE comparison the checklist itself describes).
    win_rate_stability = None
    if len(closed) >= 4:
        core_cut = int(len(closed) * 0.7)
        core_early, core_late = closed[:core_cut], closed[core_cut:]
        early_wr = win_rate([t["Return"] for t in core_early])
        late_wr = win_rate([t["Return"] for t in core_late])
        wr_diff = late_wr - early_wr
        # Verdict withheld (numbers still exposed) below ~5 late-window
        # trades -- a 2-trade late set (n=5 total) can swing 66pp on one
        # trade, which isn't a real "fading" signal (found in paired review
        # 2026-08-09, same spirit as overlay_robustness's len(half)>=2 guard).
        win_rate_stability = {
            "early_win_rate_pct": early_wr, "late_win_rate_pct": late_wr, "diff_pct": wr_diff,
            "verdict": (None if len(core_late) < 5 else ("FADING" if wr_diff < -20 else "STABLE")),
        }

    # Checklist items 9/10: how much of this node's edge depends on capital a
    # real cash-account same-day-re-buy rule (schwab_safety.py's same-day-
    # block, enforced live) would actually block. TrailingBoth-only -- the
    # checklist itself scopes this there, and run_backtest_v18 (TrailingExit)
    # has no same_day_block parameter to even test. One extra kernel run
    # (blocked variant), reusing the same prep/node params as the unblocked
    # run above; #10 (stability) reuses the SAME two trade lists, chronologically
    # split, rather than a third kernel run.
    same_day_block_check = None
    if strategy == "TrailingBothZScoreBreakout" and closed:
        # return_bounds=True so this compares ROBUST alpha (MIN of possible/
        # pessimistic/certain) on both sides, matching the checklist's own
        # explicit instruction ("compare both trade count and robust alpha,
        # not just one or the other") -- an earlier version compared
        # possible-only, found wrong in paired review 2026-08-09.
        t_block, p_block, c_block = run_backtest_v110(
            df_hourly=None, df_daily_indicators=None, ticker=ticker,
            take_profit=node["arm_pct"] / 100.0, stop_loss=node["sl"] / 100.0,
            max_hours_to_hold=node["hold"], z_score_threshold=node["z"],
            trail_buy_pct=node["trail_buy_pct"] / 100.0, trail_pct=node["trail_sell_pct"] / 100.0,
            entry_timing=node["entry_timing"], return_bounds=True, prep=prep, same_day_block=True,
        )
        closed_block = [t for t in t_block if t["Result"] in ("WIN", "LOSS", "TWIN", "TLOSS")]
        if closed_block:
            closed_pess = [t for t in p_base if t["Result"] in ("WIN", "LOSS", "TWIN", "TLOSS")] if p_base else []
            closed_cert = [t for t in c_base if t["Result"] in ("WIN", "LOSS", "TWIN", "TLOSS")] if c_base else []
            closed_block_pess = [t for t in p_block if t["Result"] in ("WIN", "LOSS", "TWIN", "TLOSS")] if p_block else []
            closed_block_cert = [t for t in c_block if t["Result"] in ("WIN", "LOSS", "TWIN", "TLOSS")] if c_block else []

            def _robust_alpha(base, pess, cert, spy_bh):
                if not base:
                    return None
                alphas = [_summarize_trades(base, spy_bh)[0]]
                if pess:
                    alphas.append(_summarize_trades(pess, spy_bh)[0])
                if cert:
                    alphas.append(_summarize_trades(cert, spy_bh)[0])
                return min(alphas)

            spy_bh_full = period_spy_bh(data_start, data_end)
            alpha_unblocked = _robust_alpha(closed, closed_pess, closed_cert, spy_bh_full)
            alpha_blocked = _robust_alpha(closed_block, closed_block_pess, closed_block_cert, spy_bh_full)
            n_unblocked, n_blocked = len(closed), len(closed_block)
            trade_retention_pct = (n_blocked / n_unblocked * 100) if n_unblocked else None
            # Only a meaningful ratio when the unblocked baseline is a real
            # positive edge -- with a negative denominator the ratio's sign
            # inverts (a blocked variant that's LESS bad would read as "low
            # retention", the opposite of what the number is meant to convey).
            # Raw before/after values are always exposed too so the sign is
            # visible even when the ratio itself is withheld (found in review).
            alpha_retention_pct = ((alpha_blocked / alpha_unblocked * 100)
                                    if (alpha_unblocked is not None and alpha_unblocked > 0
                                        and alpha_blocked is not None) else None)

            # Stability (#10): split at ONE shared calendar date (the
            # unblocked run's own 70th-percentile trade's entry time) so both
            # variants are compared over the SAME real time window -- an
            # earlier version cut each list at its own 70%-of-COUNT index,
            # which lands at two different dates once same-day-block removes
            # trades non-uniformly in time (found in review). Per-half SPY
            # benchmark, not the full-window constant, matching
            # train_test_split_check.py's own established convention.
            def _filter_period(trades, start, end):
                return [t for t in trades if start <= t["Entry Time"] < end]

            cut_idx = max(0, min(len(closed) - 1, int(len(closed) * 0.7)))
            cut_time = closed[cut_idx]["Entry Time"] if closed else None
            retention_early = retention_late = None
            if cut_time is not None:
                spy_early = period_spy_bh(data_start, cut_time)
                spy_late = period_spy_bh(cut_time, data_end)
                a_u_early = _robust_alpha(_filter_period(closed, data_start, cut_time),
                                           _filter_period(closed_pess, data_start, cut_time),
                                           _filter_period(closed_cert, data_start, cut_time), spy_early)
                a_b_early = _robust_alpha(_filter_period(closed_block, data_start, cut_time),
                                           _filter_period(closed_block_pess, data_start, cut_time),
                                           _filter_period(closed_block_cert, data_start, cut_time), spy_early)
                a_u_late = _robust_alpha(_filter_period(closed, cut_time, data_end),
                                          _filter_period(closed_pess, cut_time, data_end),
                                          _filter_period(closed_cert, cut_time, data_end), spy_late)
                a_b_late = _robust_alpha(_filter_period(closed_block, cut_time, data_end),
                                          _filter_period(closed_block_pess, cut_time, data_end),
                                          _filter_period(closed_block_cert, cut_time, data_end), spy_late)
                if a_u_early is not None and a_u_early > 0:
                    retention_early = (a_b_early / a_u_early * 100) if a_b_early is not None else 0.0
                if a_u_late is not None and a_u_late > 0:
                    retention_late = (a_b_late / a_u_late * 100) if a_b_late is not None else 0.0

            # Raw compounded-return pair (2026-08-12), added alongside the
            # existing alpha-based retention numbers above so same_day_block
            # gets the same "with-lever vs without-lever compounded return"
            # treatment drought/addon already have (drought_compounded_pct/
            # addon_compounded_pct) -- previously same_day_block only
            # exposed an alpha RATIO, never the actual before/after % return
            # this report's stacked-CAGR columns (core_addon_cagr_pct etc.)
            # need to fold it in the same way.
            compounded_unblocked_pct = compounded([t["Return"] for t in closed])
            compounded_blocked_pct = compounded([t["Return"] for t in closed_block])
            same_day_block_check = {
                "trade_retention_pct": trade_retention_pct,
                "alpha_retention_pct": alpha_retention_pct,
                "alpha_unblocked_pct": alpha_unblocked,
                "alpha_blocked_pct": alpha_blocked,
                "retention_early_pct": retention_early,
                "retention_late_pct": retention_late,
                "compounded_unblocked_pct": compounded_unblocked_pct,
                "compounded_blocked_pct": compounded_blocked_pct,
            }

    current_dd_pct = None
    if closed:
        equity, peak = 100.0, 100.0
        for t in closed:
            equity *= (1.0 + t["Return"])
            peak = max(peak, equity)
        current_dd_pct = (equity - peak) / peak * 100.0

    return {
        "folds_with_trades": fwt,
        "positive_folds": positive_folds,
        "total_folds": n_folds,
        "min_fold_alpha": summary["min_fold_alpha"],
        "max_fold_alpha": summary["max_fold_alpha"],
        "mean_fold_alpha": summary["mean_fold_alpha"],
        "verdict": verdict,
        "max_drawdown_pct": max_dd_pct,
        "current_drawdown_pct": current_dd_pct,
        "same_day_block_check": same_day_block_check,
        "fluke_check": fluke_check,
        "win_rate_stability": win_rate_stability,
        "crash_2025_check": crash_2025_check,
    }


# K-1/UBTI tax-filer status -- ONLY reflects tickers actually confirmed via a
# real source check (issuer tax-document page), per this project's standing
# "confirmed via real search, not assumed" convention (see docs/research_log.md's
# 2026-08-07 tax/compliance entry, TQQQ 2026-08-04 entry). Every ticker not
# listed here is genuinely unchecked -- default to "not checked", never guess
# from structural resemblance to a confirmed ticker. Add a row here only after
# a real source check, with the source noted.
K1_STATUS = {
    "AGQ": "CONFIRMED K-1 (proshares.com tax-and-filing-documents)",
    "USO": "CONFIRMED K-1 (uscfinvestments.com/tax-information)",
    "UCO": "CONFIRMED K-1 (same ProShares commodity-pool family as USO)",
    "SCO": "CONFIRMED K-1 (same ProShares commodity-pool family as USO)",
    "ZSL": "CONFIRMED K-1 (same ProShares silver-futures family as AGQ)",
    "UVIX": "CONFIRMED K-1 (VIX-futures-based, taxpackagesupport.com/volatility_uvix)",
    "SOXL": "confirmed clean, standard 1099 (direxion.com FAQ)",
    "NUGT": "confirmed clean, standard 1099 (direxion.com FAQ)",
    "TQQQ": "confirmed clean, standard 1099 (RIC, not a commodity/currency/volatility product)",
    "GDXU": "not K-1, but a BMO-issued ETN (unsecured issuer/counterparty credit risk) -- bmoetns.com",
    "GDXD": "not K-1, but a BMO-issued ETN (unsecured issuer/counterparty credit risk) -- bmoetns.com",
}


def k1_status(conn, ticker):
    """Reads from tickers.k1_status (the real screener table) -- migrated
    2026-08-09 off the hardcoded K1_STATUS dict above, which is kept only as
    the historical seed data (see the DB populate step in session history)."""
    row = conn.execute("SELECT k1_status FROM tickers WHERE symbol=?", (ticker,)).fetchone()
    return row[0] if row and row[0] else "not checked"


def underlier_info(conn, ticker):
    """(underlier_count, underlier_note) from tickers -- the same real
    screener table k1_status now reads from. Both None if never researched."""
    row = conn.execute("SELECT underlier_count, underlier_note FROM tickers WHERE symbol=?", (ticker,)).fetchone()
    return (row[0], row[1]) if row else (None, None)


_BEAR_UNDERLYING_CACHE = {}


def bear_market_stress_check(ticker, strategy, node):
    """Real historical-crash stress test (scripts/sim_bear_market_stress.py),
    reused directly rather than reimplemented -- synthesizes a daily-bar
    price series for `ticker` from a long-history 1x underlying proxy
    (yfinance back to 2000) and runs a daily-bar approximation of THIS row's
    exact node params through each of the 4 real historical crash windows
    (2008 GFC, 2020 COVID, 2022 bear, 2000 dotcom -- sim_bear_market_stress.CRASHES).

    Differs from the original script's own CLI in one respect: that script
    only ever looks up a ticker's real watch_list node (get_node_params);
    this passes the row's OWN candidate node params directly, so a
    'best possible'/'best certain'/etc. candidate that hasn't been promoted
    to watch_list yet still gets tested, not just the currently-live config.

    Returns None if `ticker` has no proxy mapping (scripts/sim_bear_market_stress.PROXIES) --
    real, not every candidate ticker has a viable long-history proxy, don't
    guess one. Else a dict:
      {proxy, leverage, note, crashes: {crash_name: {
         decline_trades, decline_compounded_pct,
         combined_trades, combined_compounded_pct, combined_max_drawdown_pct, combined_win_rate_pct,
         buy_hold_return_pct, truncated_start (real vs official crash-window start)}},
       worst_crash, worst_compounded_return_pct}

    Reports the DECLINE-leg return separately from the combined (decline+
    recovery) return (found in paired review 2026-08-09: the original
    version only kept "combined," but "does this survive a crash" is a
    decline-leg question and can diverge in sign from the combined number --
    e.g. a real run showed 2022_bear decline +0.8% vs combined -13.1%).
    combined_max_drawdown_pct is exposed too (previously computed then
    discarded) -- the single most decision-relevant number for a promotion
    call. A 0-trade crash is NOT scored/bucketed as a "win" -- see
    `worst_crash` selection below, and `combined_trades`/`decline_trades`
    are always present so a reader can tell "never triggered here" apart
    from "traded and finished flat."
    """
    if ticker not in _BEAR_PROXIES:
        return None
    proxy, leverage, note = _BEAR_PROXIES[ticker]

    if proxy not in _BEAR_UNDERLYING_CACHE:
        try:
            _BEAR_UNDERLYING_CACHE[proxy] = _bear_fetch_underlying(proxy)
        except Exception as e:
            print(f"bear_market_stress_check({ticker}): fetch_underlying({proxy}) failed: {e}", file=sys.stderr)
            _BEAR_UNDERLYING_CACHE[proxy] = None
    underlying = _BEAR_UNDERLYING_CACHE[proxy]
    if underlying is None:
        return None
    synth = _bear_synthesize(underlying, leverage)

    params = {
        "strategy": strategy, "window": node["window"], "z_score_threshold": node["z"],
        "fixed_sl": node["sl"], "trail_buy_pct": node["trail_buy_pct"], "trail_sell_pct": node["trail_sell_pct"],
        "arm_sell_pct": node["arm_pct"], "take_profit": None, "max_hold_hours": node["hold"],
    }

    crash_results = {}
    for crash, (start, bottom, recov_end) in _BEAR_CRASHES.items():
        start_ts, bottom_ts, end_ts = _pd.Timestamp(start), _pd.Timestamp(bottom), _pd.Timestamp(recov_end)
        window = synth.loc[start:recov_end]
        if len(window) < node["window"] + 5:
            crash_results[crash] = {"combined_trades": 0, "note": "insufficient data in window"}
            continue
        # Real proxy history can start AFTER the crash's official decline
        # leg (e.g. SOXX launched 2001-07, after the 2000-03 dotcom peak) --
        # flagged here rather than silently presenting a truncated window as
        # the full crash (found in paired review 2026-08-09: this was
        # previously unflagged despite the source script's own comment
        # claiming it would be).
        truncated_start = bool(underlying.index.min() > start_ts)
        lookback_start = synth.index.searchsorted(start_ts) - node["window"] - 1
        sim_bars = synth.iloc[max(0, lookback_start):synth.index.searchsorted(end_ts) + 1]
        all_trades = _bear_run_strategy(sim_bars, params)
        win_start_i = sim_bars.index.searchsorted(start_ts)
        bottom_i = sim_bars.index.searchsorted(bottom_ts)
        all_trades = [t for t in all_trades if t["entry_day"] >= win_start_i]
        decline_trades = [t for t in all_trades if t["entry_day"] < bottom_i]

        combined_summary = _bear_summarize(all_trades, sim_bars, start_ts, end_ts)
        decline_summary = _bear_summarize(decline_trades, sim_bars, start_ts, bottom_ts)
        crash_results[crash] = {
            "decline_trades": decline_summary["trades"],
            "decline_compounded_pct": decline_summary["compounded_return"] * 100,
            "combined_trades": combined_summary["trades"],
            "combined_compounded_pct": combined_summary["compounded_return"] * 100,
            "combined_max_drawdown_pct": (combined_summary["max_drawdown"] * 100
                                           if combined_summary["max_drawdown"] is not None else None),
            "combined_win_rate_pct": (combined_summary["win_rate"] * 100
                                       if combined_summary["win_rate"] is not None else None),
            "buy_hold_return_pct": combined_summary["buy_hold_return"] * 100,
            "truncated_start": truncated_start,
        }

    # A 0-trade crash reports compounded_pct=0.0 by construction (see
    # sim_bear_market_stress.summarize) -- excluded from "worst crash"
    # scoring so a node that simply never triggered during some window
    # can't win/mask the comparison as if it had survived a real crash
    # unscathed (found in paired review 2026-08-09).
    scored = {c: r["combined_compounded_pct"] for c, r in crash_results.items()
              if r.get("combined_trades", 0) > 0}
    worst_crash = min(scored, key=scored.get) if scored else None
    return {
        "proxy": proxy, "leverage": leverage, "note": note, "crashes": crash_results,
        "worst_crash": worst_crash,
        "worst_compounded_return_pct": scored.get(worst_crash) if worst_crash else None,
    }


def ticker_sector(conn, ticker):
    """Real description/leverage/inverse straight from the tickers table --
    not a hand-typed lookup."""
    c = conn.cursor()
    c.execute("SELECT description, leverage, inverse FROM tickers WHERE symbol=?", (ticker,))
    row = c.fetchone()
    if not row:
        return None
    desc, leverage, inverse = row
    lev_str = f"{leverage:g}x" if leverage else ""
    inv_str = " inverse" if inverse else ""
    return f"{desc} ({lev_str}{inv_str})" if desc else None


def _bucket(value, edges, labels):
    """edges is a sorted list of upper bounds; labels has len(edges)+1 entries."""
    if value is None:
        return None
    for edge, label in zip(edges, labels):
        if value < edge:
            return label
    return labels[-1]


def add_tranches(rec):
    """Adds clean, pivot-friendly categorical columns for every continuous/
    free-text criterion already in `rec` -- per the user's explicit 2026-08-09
    call: 'give me a summary column for each criteria I can use in a pivot
    table... create tranches for anything not immediately pass/fail.' Mutates
    and returns rec. Every _tranche value is indicative only, not a hard gate
    (some of these are harder stops than others -- left to the user to decide
    which, via filtering the sheet themselves)."""
    rec["years_tranche"] = _bucket(rec.get("years"), [1, 2, 3], ["<1y", "1-2y", "2-3y", "3y+"])
    rec["trades_tranche"] = _bucket(rec.get("trades"), [10, 25, 50, 100],
                                     ["<10", "10-25", "25-50", "50-100", "100+"])
    rec["liquidity_tranche"] = _bucket(rec.get("liquidity_dollars_per_day"), [50_000, 100_000, 500_000, 1_000_000],
                                        ["<$50k", "$50k-100k", "$100k-500k", "$500k-1M", "$1M+"])
    cagr_val = rec.get("strategy_cagr_pct")
    rec["cagr_tranche"] = ("negative" if cagr_val is not None and cagr_val < 0 else
                            _bucket(cagr_val, [50, 100, 200], ["0-50%", "50-100%", "100-200%", "200%+"]))
    rec["cliff_tranche"] = rec.get("status")  # SAFE/CLIFF -- clean alias, consistent _tranche naming

    wf = rec.get("walk_forward")
    if not wf:
        rec["wf_tranche"] = "NONE"
    elif wf["verdict"].startswith("THIN"):
        rec["wf_tranche"] = "THIN"
    elif wf["verdict"].startswith("PASS"):
        rec["wf_tranche"] = "PASS"
    elif wf["verdict"].startswith("MARGINAL"):
        rec["wf_tranche"] = "MARGINAL"
    else:
        rec["wf_tranche"] = "FAIL"

    def _overlay_tranche(key):
        r = rec.get(key)
        if not r:
            return "NONE"
        return "OK" if r["verdict"] == "OK" else "FRAGILE"
    rec["addon_tranche"] = _overlay_tranche("addon_robustness")
    rec["drought_tranche"] = _overlay_tranche("drought_robustness")

    def _wr_tranche(r):
        if not r or r.get("win_rate_verdict") is None:
            return "NONE"
        return r["win_rate_verdict"]
    rec["addon_wr_tranche"] = _wr_tranche(rec.get("addon_robustness"))
    rec["drought_wr_tranche"] = _wr_tranche(rec.get("drought_robustness"))

    ie = rec.get("drought_included_excluded")
    if not ie:
        rec["drought_ie_tranche"] = "NONE"
    elif ie["verdict"].startswith("N/A"):
        rec["drought_ie_tranche"] = "N/A"
    else:
        rec["drought_ie_tranche"] = ie["verdict"]

    bm = rec.get("bear_market")
    if not bm:
        rec["bear_market_tranche"] = "NO_PROXY" if rec.get("ticker") not in _BEAR_PROXIES else "NONE"
    else:
        worst = bm.get("worst_compounded_return_pct")
        # worst_crash/worst is None specifically when EVERY crash window had
        # zero trades (bear_market_stress_check excludes 0-trade crashes
        # from scoring) -- a real, distinct case from "traded and broke
        # even," not the same as POSITIVE (found in paired review 2026-08-09).
        if worst is None:
            rec["bear_market_tranche"] = "NO_TRADES_ANY_CRASH"
        elif worst >= 0:
            rec["bear_market_tranche"] = "POSITIVE"
        else:
            rec["bear_market_tranche"] = "LOSS_" + _bucket(-worst, [10, 50, 90],
                                                             ["<10%", "10-50%", "50-90%", "90%+"])

    k1 = rec.get("k1_status", "not checked")
    if k1.startswith("CONFIRMED K-1"):
        rec["k1_tranche"] = "K1_CONFIRMED"
    elif k1.startswith("confirmed clean"):
        rec["k1_tranche"] = "CLEAN_CONFIRMED"
    elif "ETN" in k1:
        rec["k1_tranche"] = "ETN_NOT_K1"
    else:
        rec["k1_tranche"] = "NOT_CHECKED"

    poss = rec.get("alpha_possible_pct")
    cert = rec.get("alpha_certain_pct")
    if poss is None or cert is None or poss == 0:
        rec["resolution_spread_tranche"] = None
    else:
        spread_pct = abs(cert - poss) / abs(poss) * 100
        rec["resolution_spread_tranche"] = _bucket(spread_pct, [20, 100], ["TIGHT", "MODERATE", "WIDE"])

    wf_rec = rec.get("walk_forward")

    dd = wf_rec.get("max_drawdown_pct") if wf_rec else None
    rec["drawdown_tranche"] = _bucket(-dd if dd is not None else None, [10, 25, 50],
                                       ["<10%", "10-25%", "25-50%", "50%+"]) if dd is not None else None

    t30 = rec.get("trend_30d_pct")
    if t30 is None:
        rec["trend_tranche"] = None
    elif t30 >= 20:
        rec["trend_tranche"] = "STRONG_UP"
    elif t30 >= 5:
        rec["trend_tranche"] = "MILD_UP"
    elif t30 > -5:
        rec["trend_tranche"] = "FLAT"
    elif t30 > -20:
        rec["trend_tranche"] = "MILD_DOWN"
    else:
        rec["trend_tranche"] = "STRONG_DOWN"

    sf = rec.get("split_flag")
    rec["split_tranche"] = None if sf is None else ("NONE" if sf == "none" else "FLAGGED")

    fluke_rec = wf_rec.get("fluke_check") if wf_rec else None
    rec["core_fluke_tranche"] = fluke_rec["verdict"] if fluke_rec else "NONE"

    core_wrs_rec = wf_rec.get("win_rate_stability") if wf_rec else None
    if not core_wrs_rec:
        rec["core_wr_tranche"] = "NONE"
    elif core_wrs_rec["verdict"] is None:
        rec["core_wr_tranche"] = "THIN_SAMPLE"
    else:
        rec["core_wr_tranche"] = core_wrs_rec["verdict"]

    crash_rec = wf_rec.get("crash_2025_check") if wf_rec else None
    rec["crash_2025_tranche"] = crash_rec["verdict"] if crash_rec else "NONE"

    sdb_rec = wf_rec.get("same_day_block_check") if wf_rec else None
    sdb_alpha_ret = sdb_rec.get("alpha_retention_pct") if sdb_rec else None
    if sdb_alpha_ret is None:
        rec["sdb_tranche"] = "N/A (TrailingExit)" if rec.get("strategy") == "TrailingExitZScoreBreakout" else None
        rec["sdb_recommend"] = rec["sdb_tranche"]
    else:
        rec["sdb_tranche"] = _bucket(sdb_alpha_ret, [25, 75], ["LOW_RETENTION", "MODERATE_RETENTION",
                                                                "HIGH_RETENTION"])
        if sdb_alpha_ret > 105:
            early, late = sdb_rec.get("retention_early_pct"), sdb_rec.get("retention_late_pct")
            # "Holds up" means the block also helps (>100%) in BOTH stability
            # halves independently -- a real, consistent gain (the LABU
            # shape), not just one lucky half dragging the overall ratio up.
            # A missing half (too few trades to split) is treated as unknown,
            # not confirmed -- stays UNSTABLE rather than defaulting to HELPS.
            if early is not None and late is not None and early > 100 and late > 100:
                rec["sdb_recommend"] = "BLOCK_HELPS"
            else:
                rec["sdb_recommend"] = "BLOCK_HELPS_UNSTABLE"
        elif sdb_alpha_ret >= 95:
            rec["sdb_recommend"] = "NEUTRAL"
        else:
            rec["sdb_recommend"] = "COSTS"
    return rec


def _build_output_row(rec):
    """Flattens one candidate_full_review record (nested dicts like
    walk_forward/addon_robustness/bear_market) into a single flat row keyed
    exactly by FIELDNAMES -- shared by both --csv and --xlsx so the two
    outputs can never drift apart."""
    if rec.get("no_data"):
        return {"ticker": rec["ticker"]}

    row = {k: rec.get(k) for k in FIELDNAMES if k in rec}
    row["ticker"] = rec["ticker"]
    row["also_matches"] = " = ".join(rec.get("also_matches", [rec["candidate_type"]]))
    ar = rec.get("addon_robustness")
    dr = rec.get("drought_robustness")
    row["addon_robustness_verdict"] = ar["verdict"] if ar else None
    row["drought_robustness_verdict"] = dr["verdict"] if dr else None
    row["addon_early_wr_pct"] = ar["early_win_rate_pct"] if ar else None
    row["addon_late_wr_pct"] = ar["late_win_rate_pct"] if ar else None
    row["addon_wr_verdict"] = ar["win_rate_verdict"] if ar else None
    row["drought_early_wr_pct"] = dr["early_win_rate_pct"] if dr else None
    row["drought_late_wr_pct"] = dr["late_win_rate_pct"] if dr else None
    row["drought_wr_verdict"] = dr["win_rate_verdict"] if dr else None

    ie = rec.get("drought_included_excluded")
    if ie:
        row["drought_ie_confirm_days"] = ie.get("confirm_days")
        row["drought_ie_vol_gate"] = ie.get("vol_gate")
        row["drought_ie_n_included"] = ie.get("n_included")
        row["drought_ie_included_compounded_pct"] = ie.get("included_compounded_pct")
        row["drought_ie_included_win_rate_pct"] = ie.get("included_win_rate_pct")
        row["drought_ie_n_excluded"] = ie.get("n_excluded")
        row["drought_ie_excluded_compounded_pct"] = ie.get("excluded_compounded_pct")
        row["drought_ie_excluded_win_rate_pct"] = ie.get("excluded_win_rate_pct")
        row["drought_ie_verdict"] = ie.get("verdict")

    wf = rec.get("walk_forward")
    fluke = wf.get("fluke_check") if wf else None
    if fluke:
        row["core_fluke_trades"] = fluke["trades"]
        row["core_fluke_alpha_pct"] = fluke["alpha_pct"]
        row["core_fluke_alpha_without_biggest_pct"] = fluke["alpha_without_biggest_pct"]
        row["core_fluke_compounded_pct"] = fluke["compounded_pct"]
        row["core_fluke_compounded_without_biggest_pct"] = fluke["compounded_without_biggest_pct"]
        row["core_fluke_verdict"] = fluke["verdict"]
    core_wrs = wf.get("win_rate_stability") if wf else None
    if core_wrs:
        row["core_wr_early_pct"] = core_wrs["early_win_rate_pct"]
        row["core_wr_late_pct"] = core_wrs["late_win_rate_pct"]
        row["core_wr_diff_pct"] = core_wrs["diff_pct"]
        row["core_wr_verdict"] = core_wrs["verdict"]

    bm = rec.get("bear_market")
    if bm:
        row["bear_proxy"] = bm.get("proxy")
        row["bear_leverage"] = bm.get("leverage")
        row["bear_note"] = bm.get("note")
        row["bear_worst_crash"] = bm.get("worst_crash")
        row["bear_worst_compounded_pct"] = bm.get("worst_compounded_return_pct")
        for crash_key in ("2008_gfc", "2020_covid", "2022_bear", "2000_dotcom"):
            cr = bm.get("crashes", {}).get(crash_key)
            row[f"bear_{crash_key}_decline_pct"] = cr.get("decline_compounded_pct") if cr else None
            row[f"bear_{crash_key}_combined_pct"] = cr.get("combined_compounded_pct") if cr else None
            row[f"bear_{crash_key}_max_dd_pct"] = cr.get("combined_max_drawdown_pct") if cr else None
            row[f"bear_{crash_key}_trades"] = cr.get("combined_trades") if cr else None
        dotcom = bm.get("crashes", {}).get("2000_dotcom")
        row["bear_2000_dotcom_truncated"] = dotcom.get("truncated_start") if dotcom else None

    crash25 = wf.get("crash_2025_check") if wf else None
    if crash25:
        row["crash25_ticker_bh_decline_pct"] = crash25["ticker_bh_decline_pct"]
        row["crash25_ticker_bh_recovery_pct"] = crash25["ticker_bh_recovery_pct"]
        row["crash25_ticker_bh_combined_pct"] = crash25["ticker_bh_combined_pct"]
        row["crash25_ticker_bh_max_dd_pct"] = crash25["ticker_bh_max_dd_pct"]
        row["crash25_algo_decline_pct"] = crash25["algo_decline_pct"]
        row["crash25_algo_recovery_pct"] = crash25["algo_recovery_pct"]
        row["crash25_algo_combined_pct"] = crash25["algo_combined_pct"]
        row["crash25_algo_max_dd_pct"] = crash25["algo_max_dd_pct"]
        row["crash25_algo_trades"] = crash25["algo_trades"]
        row["crash25_verdict"] = crash25["verdict"]
    row["wf_verdict"] = wf["verdict"] if wf else None
    row["wf_positive_folds"] = wf["positive_folds"] if wf else None
    row["wf_total_folds"] = wf["total_folds"] if wf else None
    row["wf_min_fold_alpha"] = wf["min_fold_alpha"] if wf else None
    row["wf_max_fold_alpha"] = wf["max_fold_alpha"] if wf else None
    row["wf_mean_fold_alpha"] = wf["mean_fold_alpha"] if wf else None
    row["max_drawdown_pct"] = wf["max_drawdown_pct"] if wf else None
    row["current_drawdown_pct"] = wf["current_drawdown_pct"] if wf else None
    sdb = wf["same_day_block_check"] if wf else None
    row["sdb_trade_retention_pct"] = sdb["trade_retention_pct"] if sdb else None
    row["sdb_alpha_retention_pct"] = sdb["alpha_retention_pct"] if sdb else None
    row["sdb_alpha_unblocked_pct"] = sdb["alpha_unblocked_pct"] if sdb else None
    row["sdb_alpha_blocked_pct"] = sdb["alpha_blocked_pct"] if sdb else None
    row["sdb_retention_early_pct"] = sdb["retention_early_pct"] if sdb else None
    row["sdb_retention_late_pct"] = sdb["retention_late_pct"] if sdb else None
    row["sdb_compounded_unblocked_pct"] = sdb["compounded_unblocked_pct"] if sdb else None
    row["sdb_compounded_blocked_pct"] = sdb["compounded_blocked_pct"] if sdb else None
    efa = rec.get("exit_fill_acc")
    row["exit_fillacc_win_pct"] = efa[0] if efa else None
    row["exit_fillacc_mean_err_pct"] = efa[1] if efa else None
    row["exit_fillacc_n"] = efa[2] if efa else None
    return row


def _timestamped_name(name, ext):
    """Appends a run timestamp to the given base name so successive --csv/--xlsx
    runs never overwrite each other -- every run is real evidence (which
    confirm_days/node config produced which numbers), same "never delete,
    versioning exists for coexistence" convention as backtest_cache."""
    stem = name[: -len(ext)] if name.endswith(ext) else name
    return f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"


def _write_xlsx(name, csv_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path("output") / (name if name.endswith(".xlsx") else f"{name}.xlsx")
    out_path.parent.mkdir(exist_ok=True)

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Full Review"

    data_ws.append(FIELDNAMES)
    for cell in data_ws[1]:
        cell.font = Font(bold=True)
    for row in csv_rows:
        data_ws.append([row.get(c) for c in FIELDNAMES])
    data_ws.freeze_panes = "B2"
    for i, col in enumerate(FIELDNAMES, start=1):
        data_ws.column_dimensions[get_column_letter(i)].width = max(12, min(len(col) + 2, 28))

    def_ws = wb.create_sheet("Column Definitions")
    def_ws.append(["Column", "Definition"])
    for cell in def_ws[1]:
        cell.font = Font(bold=True)
    for col, definition in COLUMN_DEFS.items():
        def_ws.append([col, definition])
        def_ws.cell(row=def_ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    def_ws.column_dimensions["A"].width = 32
    def_ws.column_dimensions["B"].width = 110

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tickers", nargs="*")
    ap.add_argument("--version", default=None,
                     help="force a single version for every ticker (old behavior). Default: auto-resolve "
                          "per ticker via resolve_version() -- v5.1 when the ticker has it, else v5.")
    ap.add_argument("--db", default=DB_PATH)
    ap.add_argument("--min-alpha", type=float, default=0,
                     help="Alpha floor for the 'best safe node' cliff-safety search (default 0 -- show "
                          "every cliff-safe node regardless of magnitude, not just the >=200%% convention).")
    ap.add_argument("--skip-5min", action="store_true")
    ap.add_argument("--skip-overlay", action="store_true")
    ap.add_argument("--skip-walkforward", action="store_true",
                     help="skip walk-forward AND everything computed alongside it in the same kernel "
                          "run -- max/current drawdown, same-day-block sensitivity/stability (checklist "
                          "items 9-13 all go blank, not just 13; slower otherwise -- runs the real numba "
                          "kernel per row, not just a DB lookup)")
    ap.add_argument("--folds", type=int, default=DEFAULT_FOLDS)
    ap.add_argument("--skip-trend", action="store_true", help="skip the 30d/90d macro-trend check")
    ap.add_argument("--skip-splits", action="store_true",
                     help="skip the stock-split check (yfinance network call per ticker)")
    ap.add_argument("--vol-gate", type=float, default=DEFAULT_VOL_GATE,
                     help="entry-vol percentile gate for the drought included-vs-excluded challenge")
    ap.add_argument("--skip-bear-market", action="store_true",
                     help="skip the historical-crash stress test (real yfinance network call per "
                          "underlying proxy, cached across tickers sharing one -- e.g. SOXL/USD/SOXS all "
                          "use SOXX -- but still real network + 4-crash daily-bar sim work per row)")
    ap.add_argument("--csv", default=None)
    ap.add_argument("--xlsx", default=None,
                     help="write output/<name>.xlsx (Full Review sheet + Column Definitions glossary sheet), "
                          "same pattern as candidate_summary_report.py --xlsx")
    ap.add_argument("--set-pick", nargs=2, metavar=("NODE_ID", "yes|no"), default=None,
                     help="record a promotion decision against a real node_id (from a prior report's node_id "
                          "column) and exit -- no report is generated. Combine with --comment.")
    ap.add_argument("--comment", default=None,
                     help="free-text note to attach alongside --set-pick (or on its own, against --set-pick's "
                          "NODE_ID, to update just the comment without touching pick).")
    args = ap.parse_args()

    if args.set_pick:
        node_id, pick = args.set_pick
        conn = sqlite3.connect(args.db)
        ensure_candidate_nodes_table(conn)
        set_pick_comment(conn, int(node_id), pick=pick, comment=args.comment)
        print(f"node_id={node_id}: pick={pick!r} comment={args.comment!r}")
        conn.close()
        return

    conn = sqlite3.connect(args.db)
    ensure_candidate_nodes_table(conn)
    ensure_overlay_table(conn)
    tickers = args.tickers
    if not tickers:
        c = conn.cursor()
        c.execute("SELECT DISTINCT ticker FROM candidate_nodes")
        tickers = [r[0] for r in c.fetchall()]

    label_to_clean = CANDIDATE_LABELS  # raw find_candidates() label -> clean UI label
    clean_to_label = {v: k for k, v in label_to_clean.items()}

    out_rows = []
    for ticker in tickers:
        version = args.version or resolve_version(conn, ticker)
        best = best_node_strategy(conn, ticker, version)
        membership = {}
        label_to_node = {}
        if best is not None:
            strategy0, _ = best
            df_t = load_ticker_df(conn, ticker, version, strategy0)
            if not df_t.empty:
                membership, label_to_node = candidate_type_membership(df_t, args.min_alpha)

        trend_30d, trend_90d = (None, None) if args.skip_trend else ticker_trend(ticker)
        split_flag = None if args.skip_splits else ticker_split_flag(ticker)

        for row in build_rows_for_ticker(conn, ticker, version, args.min_alpha, args.skip_5min,
                                          args.skip_overlay):
            if row[1] is None:
                out_rows.append({"ticker": ticker, "no_data": True})
                continue
            rec = _row_to_record(row)
            years = rec["years"]
            strat_cagr = cagr(rec["abs_return_pct"], years * 365.25) if years else None
            rec["strategy_cagr_pct"] = strat_cagr
            rec["sector"] = ticker_sector(conn, ticker)
            rec["k1_status"] = k1_status(conn, ticker)
            rec["underlier_count"], rec["underlier_note"] = underlier_info(conn, ticker)
            rec["trend_30d_pct"] = trend_30d
            rec["trend_90d_pct"] = trend_90d
            rec["split_flag"] = split_flag

            (t, candidate_type, strategy, ralpha, sret, yrs, trades, ann_excess, fill_acc, wn, cliff,
             addon, drought, addon_mult, drought_mult, liquidity) = row
            # Look this row's exact node params up directly from label_to_node
            # (computed fresh, same session, same function find_candidates()
            # itself uses) instead of round-tripping through a DB query keyed
            # on `robust_alpha` float equality -- that query could silently
            # return no match after any backtest_cache recompute, or match a
            # DIFFERENT node on a robust_alpha tie between two grid plateaus
            # (found 2026-08-09, paired review). Also means the three raw
            # fill-resolution alphas (alpha_raw/alpha_pessimistic/alpha_certain
            # -- already present on every node dict find_candidates() produces)
            # no longer need a second DB query either.
            raw_label = clean_to_label.get(candidate_type)
            node = label_to_node.get(raw_label) if raw_label else None
            if node:
                node_id = get_or_create_candidate_node(conn, {
                    "ticker": ticker, "strategy": strategy, "version": version,
                    "window": node["window"], "z": node["z"], "fixed_sl": node["sl"], "arm_pct": node["arm_pct"],
                    "trail_buy_pct": node["trail_buy_pct"], "trail_sell_pct": node["trail_sell_pct"],
                    "max_hold_hours": node["hold"], "entry_timing": node["entry_timing"],
                    "robust_alpha": node["robust_alpha"], "trades": node["trades"],
                    "sweep_run_id": node.get("sweep_run_id"),
                })
                rec["node_id"] = node_id
                rec["pick"], rec["comment"] = get_pick_comment(conn, node_id)
                rec["addon_robustness"] = overlay_robustness(conn, ticker, strategy, version, "addon", node)
                rec["drought_robustness"] = overlay_robustness(conn, ticker, strategy, version, "drought", node)
                core_factor = 1.0 + sret / 100.0
                addon_ok = rec["addon_robustness"] is not None and rec["addon_robustness"]["verdict"] == "OK"
                drought_ok = rec["drought_robustness"] is not None and rec["drought_robustness"]["verdict"] == "OK"
                addon_factor_gated = (1.0 + addon[1] / 100.0) if (addon and addon_ok) else 1.0
                drought_factor_gated = (1.0 + drought[1] / 100.0) if (drought and drought_ok) else 1.0
                rec["drought_included_excluded"] = None if args.skip_overlay else drought_included_excluded_check(
                    conn, ticker, strategy, version, node, vol_gate=args.vol_gate)
                # If the vol-gate is confirmed to do real differential
                # selection (not just look profitable in isolation), its
                # included-only return REPLACES the plain drought factor for
                # the stacked CAGR columns -- the ungated drought_ok/
                # drought_factor_gated above is a strictly worse estimate
                # once a validated filter exists for this node (found
                # 2026-08-09: the two computations were previously silently
                # disconnected -- drought_ie_* was diagnostic-only and never
                # fed back into +Drght%/+Both%).
                ie = rec["drought_included_excluded"]
                if ie is not None and ie.get("verdict") == "REAL_SELECTION":
                    drought_factor_gated = 1.0 + ie["included_compounded_pct"] / 100.0
                days_span = years * 365.25 if years else None
                rec["core_addon_cagr_pct"] = cagr((core_factor * addon_factor_gated - 1.0) * 100.0, days_span)
                rec["core_drought_cagr_pct"] = cagr((core_factor * drought_factor_gated - 1.0) * 100.0, days_span)
                rec["core_both_cagr_pct"] = cagr(
                    (core_factor * addon_factor_gated * drought_factor_gated - 1.0) * 100.0, days_span)
                node_key = _node_key(node)
                rec["also_matches"] = membership.get(node_key, [candidate_type])
                rec["alpha_possible_pct"] = node["alpha_raw"]
                rec["alpha_pessimistic_pct"] = (node["alpha_pessimistic"] if node["alpha_pessimistic"] is not None
                                                 else node["alpha_raw"])
                rec["alpha_certain_pct"] = (node["alpha_certain"] if node["alpha_certain"] is not None
                                             else node["alpha_raw"])
                rec["walk_forward"] = None if args.skip_walkforward else core_walk_forward(
                    ticker, strategy, node, args.folds)
                sdb_check = rec["walk_forward"]["same_day_block_check"] if rec["walk_forward"] else None
                rec["core_sdb_cagr_pct"] = (cagr(sdb_check["compounded_blocked_pct"], days_span)
                                             if sdb_check and sdb_check.get("compounded_blocked_pct") is not None
                                             else None)
                rec["exit_fill_acc"] = None if args.skip_5min else exit_fill_accuracy_summary(
                    ticker, strategy, node)
                rec["bear_market"] = None if args.skip_bear_market else bear_market_stress_check(
                    ticker, strategy, node)
            else:
                rec["addon_robustness"] = None
                rec["drought_robustness"] = None
                rec["core_addon_cagr_pct"] = None
                rec["core_drought_cagr_pct"] = None
                rec["core_both_cagr_pct"] = None
                rec["core_sdb_cagr_pct"] = None
                rec["drought_included_excluded"] = None
                rec["bear_market"] = None
                rec["also_matches"] = [candidate_type]
                rec["alpha_possible_pct"] = rec["alpha_pessimistic_pct"] = rec["alpha_certain_pct"] = None
                rec["walk_forward"] = None
                rec["exit_fill_acc"] = None
            add_tranches(rec)
            out_rows.append(rec)

    conn.close()

    if args.csv or args.xlsx:
        csv_rows = [_build_output_row(rec) for rec in out_rows]

    if args.csv:
        csv_name = _timestamped_name(args.csv, ".csv")
        out_path = Path("output") / csv_name
        out_path.parent.mkdir(exist_ok=True)
        with open(out_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=FIELDNAMES)
            w.writeheader()
            for row in csv_rows:
                w.writerow(row)
        print(f"Wrote {out_path} ({len(out_rows)} rows)")

    if args.xlsx:
        xlsx_name = _timestamped_name(args.xlsx, ".xlsx")
        _write_xlsx(xlsx_name, csv_rows)
        print(f"Wrote output/{xlsx_name} ({len(out_rows)} rows)")

    if args.csv or args.xlsx:
        return

    hdr = ("%-8s %-20s %12s %8s %9s %9s %10s %6s %6s %6s %6s | %-28s | %-28s | %8s %8s %8s %8s" % (
        "Ticker", "Candidate", "Liquidity$/d", "CAGR%", "AbsRet%", "AnnExcess%", "WorstNb%",
        "Years", "Trades", "Status", "Fill", "Addon(n,comp%,WR%,robust)", "Drought(n,comp%,WR%,robust)",
        "+Addon%", "+Drght%", "+Both%", "SDB%"))
    print(hdr)
    print("-" * len(hdr))
    for rec in out_rows:
        if rec.get("no_data"):
            print(f"{rec['ticker']:8} NO_DATA")
            continue
        liq = rec['liquidity_dollars_per_day']
        liq_str = f"${liq:,.0f}" if liq is not None else "n/a"
        cagr_str = f"{rec['strategy_cagr_pct']:+.1f}" if rec['strategy_cagr_pct'] is not None else "-"
        ae_str = f"{rec['ann_excess_pct']:+.1f}" if rec['ann_excess_pct'] is not None else "-"
        wn_str = f"{rec['worst_neighbor_pct']:+.1f}" if rec['worst_neighbor_pct'] is not None else "n/a"
        fill_str = "Y" if rec['fillacc_possible_win_pct'] is not None else "-"

        def overlay_str(n_key, comp_key, wr_key, robust_key):
            n = rec.get(n_key)
            if n is None:
                return "-"
            robust = rec.get(robust_key)
            v = robust["verdict"] if robust else "n/a"
            return f"{n},{rec[comp_key]:+.1f}%,{rec[wr_key]:.0f}%,{v}"

        ao_str = overlay_str("addon_n", "addon_compounded_pct", "addon_win_rate_pct", "addon_robustness")
        dr_str = overlay_str("drought_n", "drought_compounded_pct", "drought_win_rate_pct", "drought_robustness")
        also = rec.get("also_matches", [rec["candidate_type"]])
        type_label = rec['candidate_type'] if len(also) < 2 else " = ".join(also)

        def stacked_str(key):
            v = rec.get(key)
            return f"{v:+.1f}" if v is not None else "-"

        print(f"{rec['ticker']:8} {type_label:<20} {liq_str:>12} {cagr_str:>8} "
              f"{rec['abs_return_pct']:>9.1f} {ae_str:>10} {wn_str:>10} {rec['years']!s:>6} {rec['trades']:>6} "
              f"{rec['status']:>6} {fill_str:>6} | {ao_str:<28} | {dr_str:<28} | "
              f"{stacked_str('core_addon_cagr_pct'):>8} {stacked_str('core_drought_cagr_pct'):>8} "
              f"{stacked_str('core_both_cagr_pct'):>8} {stacked_str('core_sdb_cagr_pct'):>8}")


if __name__ == "__main__":
    main()
