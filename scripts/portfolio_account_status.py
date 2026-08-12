"""Print every real watch_list node, grouped by account -- the portfolio-wide
answer to "what's running where," across every account/watchlist, not just
the active one. Unfiltered by default (every node, including legacy/stale
versions) -- use the flags below to narrow when the full dump is too much,
or export to csv/xlsx to review/clean up the full list outside the terminal.

Usage:
  .venv/bin/python scripts/portfolio_account_status.py
  .venv/bin/python scripts/portfolio_account_status.py --account ira --account soxl_ira
  .venv/bin/python scripts/portfolio_account_status.py --state live --state dry_run
  .venv/bin/python scripts/portfolio_account_status.py --strategy TrailingBoth
  .venv/bin/python scripts/portfolio_account_status.py --ticker SOXL --ticker AGQ
  .venv/bin/python scripts/portfolio_account_status.py --csv portfolio_state   # output/portfolio_state.csv
  .venv/bin/python scripts/portfolio_account_status.py --xlsx portfolio_state  # output/portfolio_state.xlsx

The terminal table shows a compact subset of columns for a quick glance;
--csv/--xlsx export EVERY watch_list column (SELECT *, not a hand-picked
subset -- a hand-picked list is exactly what silently dropped `label` the
first time this script was written) plus the derived position/account
columns below.

Reuses signals_db.get_real_position_state (the canonical node-scoped
real/paper position+pending-buy state, see its docstring) instead of
re-deriving position status by hand -- that's the exact duplication that let
the same "flat" misclassification bug recur across status_check.py/
watchlist_status.py/audit_live_test_candidates.py independently (see
docs/deep_backlog.md's 2026-08-04/2026-08-05 entries).
"""
import argparse
import csv
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import signals_db as db
import schwab_safety

RESEARCH_DB_PATH = "cache/research/trading_universe.db"

# Same param-tuple key as locate_best_node.get_or_create_candidate_node -- the
# only existing convention in this codebase for matching a watch_list node
# back to its candidate_nodes row (there's no stored FK, see the conversation
# that prompted this: watch_list nodes and candidate_nodes were never linked
# at promotion time). watch_list's z_score_threshold/arm_sell_pct map to
# candidate_nodes' z/arm_pct -- same values, different column names.
_CANDIDATE_KEY_COLS = (
    ("ticker", "ticker", str),
    ("strategy", "strategy", str),
    ("version", "version", str),
    ("window", "window", int),
    ("z_score_threshold", "z", float),
    ("fixed_sl", "fixed_sl", float),
    ("arm_sell_pct", "arm_pct", float),
    ("trail_buy_pct", "trail_buy_pct", float),
    ("trail_sell_pct", "trail_sell_pct", float),
    ("max_hold_hours", "max_hold_hours", int),
    ("entry_timing", "entry_timing", str),
)


def _candidate_key(row, wl_col_names):
    """Builds the lookup key from either a watch_list row or a candidate_nodes
    row -- wl_col_names picks which side's column names to read (watch_list's
    own vs candidate_nodes'). Returns None if any key field is missing/None
    (a node with e.g. no fixed_sl set can never match -- no partial matches)."""
    key = []
    for wl_col, cand_col, cast in _CANDIDATE_KEY_COLS:
        col = wl_col if wl_col_names else cand_col
        val = row[col]
        if val is None:
            return None
        try:
            key.append(cast(val))
        except (TypeError, ValueError):
            return None
    return tuple(key)


def _load_candidate_lookup():
    """One-time load of every candidate_nodes row into a {param_key: (id,
    robust_alpha, trades)} dict, so each watch_list row's match is an O(1)
    dict lookup instead of N queries against a second database file."""
    lookup = {}
    conn = sqlite3.connect(RESEARCH_DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute("SELECT * FROM candidate_nodes"):
            key = _candidate_key(row, wl_col_names=False)
            if key is not None and key not in lookup:
                lookup[key] = (row["id"], row["robust_alpha"], row["trades"])
    finally:
        conn.close()
    return lookup


_DERIVED_COLUMN_DEFS = {
    "trading_enabled": "schwab_safety.ACCOUNTS[account].trading_enabled -- whether this account can place real (non-dry_run) orders at all (None if account isn't in ACCOUNTS).",
    "position_status": "Real-time status from signals_db.get_real_position_state: flat / pending_entry / holding / holding_paper.",
    "position_detail": "Shares@entry_price if holding, 'resting order' if a trailing-buy is pending, else blank.",
    "candidate_node_id": "Best-effort match to candidate_nodes.id (cache/research/trading_universe.db) by exact param tuple (ticker/strategy/version/window/z/fixed_sl/arm_pct/trail_buy_pct/trail_sell_pct/max_hold_hours/entry_timing) -- same match convention as locate_best_node.get_or_create_candidate_node. No stored FK exists between the two tables, so this is a live recompute every run, not a persisted link. Blank if no candidate_nodes row has the exact same params (common for canary/soxl_test/manually-tuned nodes, which were never sourced from the candidate pipeline), or if params drifted after promotion.",
    "candidate_robust_alpha": "robust_alpha (MIN of possible/pessimistic/certain) recorded on the matched candidate_nodes row, for direct comparison against what's actually running live -- confirms the promoted node really is the validated one, not just a same-ticker lookalike.",
    "candidate_trades": "Trade count recorded on the matched candidate_nodes row.",
}

# watch_list's own column comments, for columns whose meaning isn't obvious from
# the name alone -- not exhaustive, the raw column names carry the rest.
_RAW_COLUMN_NOTES = {
    "id": "watch_list.id -- the unique node identifier (wl_id) used everywhere else in the codebase to disambiguate same-ticker nodes.",
    "watchlist_id": "Which watchlist this node belongs to (see watchlists.is_active for the current live one).",
    "version": "Backtest campaign tag or ad hoc label (canary/soxl_test/v5-overlay-test-* etc) this node's config was seeded from.",
    "label": "Short human-assigned name for this node, if set.",
    "annotation": "Longer free-text note (walk-forward status, known caveats, etc), if set.",
    "state": "One of research/paper/dry_run/live. live + trading_enabled=True on the account is real capital.",
    "entry_timing": "close (bar-close signal) or open_check (also polls the 9:31-9:40/14:31-14:40 windows).",
    "paper_role": "NULL for a normal live-track paper node, 'daily_sync' for a daily-track clone (hourly-close pricing, nightly reconciliation only).",
    "cached_avg_vol_10d": "Snapshot of 10-day average dollar volume, used for liquidity-cap checks.",
    "alpha": "Snapshot of backtest_cache.alpha_vs_spy at the time this node was added/last synced (see backfill_watch_list_alpha.py) -- not live-updated.",
}


def _build_record(r, candidate_lookup):
    pos_state = db.get_real_position_state(r["id"])
    limits = schwab_safety.ACCOUNTS.get(r["account"])
    rec = dict(r)
    rec["account"] = rec["account"] or "(none)"
    rec["trading_enabled"] = limits.trading_enabled if limits is not None else None
    rec["position_status"] = pos_state["status"]
    rec["position_detail"] = _position_detail(pos_state)
    key = _candidate_key(r, wl_col_names=True)
    match = candidate_lookup.get(key) if key is not None else None
    rec["candidate_node_id"] = match[0] if match else None
    rec["candidate_robust_alpha"] = match[1] if match else None
    rec["candidate_trades"] = match[2] if match else None
    return rec

_STRATEGY_ABBREV = {
    "TrailingBothZScoreBreakout": "TrailingBoth",
    "TrailingExitZScoreBreakout": "TrailingExit",
    "LimitOrderZScoreBreakout": "LimitOrder",
    "LimitOrderTrailingExit": "LimitOrderTE",
    "TrailingBuyZScoreBreakout": "TrailingBuy",
    "LimitExitZScoreBreakout": "LimitExit",
    "TrendFilteredZScore": "TrendFiltered",
    "ZScoreBreakout": "ZScore",
}

_STATUS_LABEL = {
    "holding": "HOLDING",
    "pending_entry": "pending",
    "holding_paper": "holding(paper)",
    "flat": "-",
}


def _trunc(s, width):
    s = s or "-"
    return s if len(s) <= width else s[: width - 1] + "…"


def _position_detail(pos_state):
    status = pos_state["status"]
    if status == "holding":
        p = pos_state["real_position"]
        return f"{p['shares']}sh@{p['entry_price']:.2f}"
    if status == "holding_paper":
        p = pos_state["paper_position"]
        return f"{p['shares']}sh@{p['entry_price']:.2f}"
    if status == "pending_entry":
        return "resting order"
    return "-"


def _timestamped_name(name, ext):
    """Appends a run timestamp so successive --csv/--xlsx runs never overwrite
    each other -- same convention as candidate_full_review.py."""
    stem = name[: -len(ext)] if name.endswith(ext) else name
    return f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"


def _write_csv(name, records, cols):
    out_path = Path("output") / _timestamped_name(name, ".csv")
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(records)
    print(f"Wrote {out_path} ({len(records)} rows, {len(cols)} columns)")


def _write_xlsx(name, records, cols, col_defs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path("output") / _timestamped_name(name, ".xlsx")
    out_path.parent.mkdir(exist_ok=True)

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Portfolio State"

    data_ws.append(cols)
    for cell in data_ws[1]:
        cell.font = Font(bold=True)
    for rec in records:
        data_ws.append([rec.get(c) for c in cols])
    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = data_ws.dimensions
    for i, col in enumerate(cols, start=1):
        data_ws.column_dimensions[get_column_letter(i)].width = max(12, min(len(col) + 2, 28))

    def_ws = wb.create_sheet("Column Definitions")
    def_ws.append(["Column", "Definition"])
    for cell in def_ws[1]:
        cell.font = Font(bold=True)
    for col in cols:
        def_ws.append([col, col_defs.get(col, "")])
        def_ws.cell(row=def_ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    def_ws.column_dimensions["A"].width = 32
    def_ws.column_dimensions["B"].width = 110

    wb.save(out_path)
    print(f"Wrote {out_path} ({len(records)} rows, {len(cols)} columns, 2 sheets)")


def _parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--account", action="append", default=None, help="filter to this account (repeatable)")
    p.add_argument("--state", action="append", default=None, help="filter to this state, e.g. live/dry_run/paper/research (repeatable)")
    p.add_argument("--strategy", action="append", default=None, help="substring match on strategy name, e.g. TrailingBoth (repeatable)")
    p.add_argument("--ticker", action="append", default=None, help="filter to this ticker (repeatable)")
    p.add_argument("--watchlist-id", type=int, action="append", default=None, help="filter to this watchlist_id (repeatable, e.g. 65 for the active 'Live v5' watchlist)")
    p.add_argument("--csv", default=None, help="write output/<name>.csv (all watch_list columns, not a subset)")
    p.add_argument("--xlsx", default=None, help="write output/<name>.xlsx (all watch_list columns + Column Definitions glossary sheet)")
    return p.parse_args()


def main():
    args = _parse_args()
    with db._conn() as c:
        raw_cols = [r[1] for r in c.execute("PRAGMA table_info(watch_list)")]
        rows = [dict(r) for r in c.execute(
            "SELECT * FROM watch_list ORDER BY account, ticker"
        ).fetchall()]

    if args.account:
        wanted = {a.lower() for a in args.account}
        rows = [r for r in rows if (r["account"] or "").lower() in wanted]
    if args.state:
        wanted = {s.lower() for s in args.state}
        rows = [r for r in rows if (r["state"] or "").lower() in wanted]
    if args.strategy:
        wanted = [s.lower() for s in args.strategy]
        rows = [r for r in rows if any(w in (r["strategy"] or "").lower() for w in wanted)]
    if args.ticker:
        wanted = {t.upper() for t in args.ticker}
        rows = [r for r in rows if r["ticker"] in wanted]
    if args.watchlist_id:
        wanted = set(args.watchlist_id)
        rows = [r for r in rows if r["watchlist_id"] in wanted]

    if args.csv or args.xlsx:
        candidate_lookup = _load_candidate_lookup()
        records = [_build_record(r, candidate_lookup) for r in rows]
        cols = raw_cols + list(_DERIVED_COLUMN_DEFS.keys())
        col_defs = {**_RAW_COLUMN_NOTES, **_DERIVED_COLUMN_DEFS}
        if args.csv:
            _write_csv(args.csv, records, cols)
        if args.xlsx:
            _write_xlsx(args.xlsx, records, cols, col_defs)
        return

    accounts = sorted({r["account"] or "(none)" for r in rows})
    for account in accounts:
        limits = schwab_safety.ACCOUNTS.get(account)
        header = f"=== {account} "
        if limits is not None:
            header += (
                f"(trading_enabled={limits.trading_enabled}, "
                f"cash_settlement_type={limits.cash_settlement_type}, "
                f"margin_capable={limits.margin_capable}, "
                f"notional_cap=${limits.notional_cap:,}) "
            )
        print(header.ljust(100, "="))
        print(f"  {'node_id':>7s}  {'ticker':6s}  {'strategy':14s} {'ver':20s} {'label':16s} "
              f"{'state':8s} {'timing':11s} {'notional':>10s} {'ovl':4s} "
              f"{'status':16s} {'position':16s}  added")

        for r in [r for r in rows if (r["account"] or "(none)") == account]:
            pos_state = db.get_real_position_state(r["id"])
            strategy = _STRATEGY_ABBREV.get(r["strategy"], r["strategy"] or "-")
            ovl = (
                ("D" if r["drought_overlay_enabled"] else "")
                + ("A" if r["addon_enabled"] else "")
                + ("S" if r["skim_enabled"] else "")
            ) or "-"
            notional = f"${r['starting_notional']:,.0f}" if r["starting_notional"] else "-"
            print(
                f"  {r['id']:>7d}  {r['ticker']:6s}  {strategy:14s} "
                f"{_trunc(r['version'], 20):20s} {_trunc(r['label'], 16):16s} "
                f"{(r['state'] or '-'):8s} "
                f"{(r['entry_timing'] or '-'):11s} {notional:>10s} {ovl:4s} "
                f"{_STATUS_LABEL[pos_state['status']]:16s} {_position_detail(pos_state):16s}  "
                f"{r['added_at']}"
            )
        print()

    print(f"{len(rows)} node(s) total across {len(accounts)} account(s).")


if __name__ == "__main__":
    main()
