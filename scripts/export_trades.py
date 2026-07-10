"""Export the full backtest trade list (every entry/exit) for a ticker to Excel,
using the same watch_list params and run_backtest_dispatch path the live sweep uses."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import sqlite3
import pandas as pd
import strategies
from backtester import prep_inputs, WIN, LOSS, TWIN, TLOSS, OPEN, _RESULT_NAMES

CACHE_DIR = Path(__file__).resolve().parent.parent / "cache"
START_CAPITAL = 50_000  # matches $50k notional per trade in live sizing


def simulate_trail_both_annotated(p, take_profit, stop_loss, max_hours_to_hold,
                                   trail_buy_pct, trail_pct, target_h0, target_h1, z_thresh):
    """Pure-Python mirror of backtester._simulate_trail_both, kept as a separate read-only
    copy (not imported from the live numba kernel) so this reporting script can record
    the extra bar indices (signal/arm) the live kernel doesn't return, without touching
    the kernel powering the live sweep engine."""
    prices, highs, lows, hours = p['prices'], p['highs'], p['lows'], p['hours']
    daily_idx, sma_arr, std_arr = p['daily_idx'], p['sma_arr'], p['std_arr']
    trend_arr, has_trend = p['trend_arr'], p['has_trend']

    trades = []
    in_trade = waiting = trailing = False
    entry_price = stop_price = tp_price = peak = 0.0
    entry_bar = held = 0
    running_low = 0.0
    wait_bars = 0
    signal_bar = None
    signal_z = None
    arm_bar = None

    n = len(prices)
    for i in range(n):
        cp, high, low = prices[i], highs[i], lows[i]

        if in_trade:
            held += 1
            if trailing:
                if high > peak:
                    peak = high
                trail_stop = peak * (1.0 - trail_pct)
                if low <= trail_stop or held >= max_hours_to_hold:
                    exit_px = trail_stop if low <= trail_stop else cp
                    pc = (exit_px - entry_price) / entry_price
                    trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                                        arm_i=arm_bar, exit_i=i, entry_p=entry_price, exit_p=exit_px,
                                        held=held, result=WIN if pc > 0 else LOSS, ret=pc))
                    in_trade = trailing = False
                continue
            if low <= stop_price:
                pc = (stop_price - entry_price) / entry_price
                trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                                    arm_i=arm_bar, exit_i=i, entry_p=entry_price, exit_p=stop_price,
                                    held=held, result=LOSS, ret=pc))
                in_trade = False
                continue
            if cp >= tp_price:
                trailing = True; peak = cp; arm_bar = i
                continue
            if held >= max_hours_to_hold:
                pc = (cp - entry_price) / entry_price
                trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                                    arm_i=arm_bar, exit_i=i, entry_p=entry_price, exit_p=cp,
                                    held=held, result=TWIN if pc > 0 else TLOSS, ret=pc))
                in_trade = False
                continue
            continue

        if waiting:
            wait_bars += 1
            if low < running_low:
                running_low = low
            buy_trigger = running_low * (1.0 + trail_buy_pct)
            if high >= buy_trigger:
                entry_price = buy_trigger
                tp_price = entry_price * (1.0 + take_profit)
                stop_price = entry_price * (1.0 - stop_loss)
                entry_bar = i; held = 0; arm_bar = None
                in_trade = True; waiting = trailing = False
                continue
            if wait_bars >= max_hours_to_hold:
                waiting = False
            continue

        h = hours[i]
        if h != target_h0 and h != target_h1:
            continue
        di = daily_idx[i]
        if di < 0:
            continue
        sma, std = sma_arr[di], std_arr[di]
        if std == 0.0:
            continue
        lower_band = sma - std * z_thresh
        signal = (cp <= lower_band) and (cp > trend_arr[di]) if has_trend else cp <= lower_band
        if signal:
            waiting = True; running_low = cp; wait_bars = 0
            signal_bar = i; signal_z = (cp - sma) / std

    if in_trade:
        cp = prices[n - 1]
        pc = (cp - entry_price) / entry_price
        trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                            arm_i=arm_bar, exit_i=n - 1, entry_p=entry_price, exit_p=cp,
                            held=held, result=OPEN, ret=pc))

    return trades


def simulate_trail_both_ohlc_aware(p, opens, take_profit, stop_loss, max_hours_to_hold,
                                    trail_buy_pct, trail_pct, target_h0, target_h1, z_thresh):
    """Side analysis only — not used by the live kernel. Same state machine as
    simulate_trail_both_annotated, except during the trailing-buy 'waiting' phase it
    resolves each bar's fill with as little guessing as possible, since Close is always
    the bar's *last* price (later than both High and Low, by definition of OHLC):

      1. CERTAIN — High clears the trigger from the *prior* bars' running_low alone
         (doesn't depend on anything that happened this bar).
      2. CERTAIN — after folding in this bar's own Low, Close itself already clears the
         new trigger. Since Close happens after the Low chronologically, this proves a
         qualifying bounce happened without needing to know exactly when.
      3. HEURISTIC — a wick clears the new trigger but Close pulls back below it before
         the bar ends: order of High vs Low within the bar is genuinely unrecoverable
         from OHLC alone, so this falls back to the Open-vs-Close bar-direction guess
         (bullish = Low-then-High, bearish = High-then-Low).
      4. NO FILL — nothing clears the trigger even after the Low update.

    Each trade record carries entry_certain so callers can report what fraction of
    fills only rely on the heuristic in case 3."""
    prices, highs, lows, hours = p['prices'], p['highs'], p['lows'], p['hours']
    daily_idx, sma_arr, std_arr = p['daily_idx'], p['sma_arr'], p['std_arr']
    trend_arr, has_trend = p['trend_arr'], p['has_trend']

    trades = []
    in_trade = waiting = trailing = False
    entry_price = stop_price = tp_price = peak = 0.0
    entry_bar = held = 0
    entry_certain = True
    running_low = 0.0
    wait_bars = 0
    signal_bar = None
    signal_z = None
    arm_bar = None

    n = len(prices)
    for i in range(n):
        cp, high, low = prices[i], highs[i], lows[i]

        if in_trade:
            held += 1
            if trailing:
                if high > peak:
                    peak = high
                trail_stop = peak * (1.0 - trail_pct)
                if low <= trail_stop or held >= max_hours_to_hold:
                    exit_px = trail_stop if low <= trail_stop else cp
                    pc = (exit_px - entry_price) / entry_price
                    trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                                        entry_certain=entry_certain, arm_i=arm_bar, exit_i=i,
                                        entry_p=entry_price, exit_p=exit_px,
                                        held=held, result=WIN if pc > 0 else LOSS, ret=pc))
                    in_trade = trailing = False
                continue
            if low <= stop_price:
                pc = (stop_price - entry_price) / entry_price
                trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                                    entry_certain=entry_certain, arm_i=arm_bar, exit_i=i,
                                    entry_p=entry_price, exit_p=stop_price,
                                    held=held, result=LOSS, ret=pc))
                in_trade = False
                continue
            if cp >= tp_price:
                trailing = True; peak = cp; arm_bar = i
                continue
            if held >= max_hours_to_hold:
                pc = (cp - entry_price) / entry_price
                trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                                    entry_certain=entry_certain, arm_i=arm_bar, exit_i=i,
                                    entry_p=entry_price, exit_p=cp,
                                    held=held, result=TWIN if pc > 0 else TLOSS, ret=pc))
                in_trade = False
                continue
            continue

        if waiting:
            wait_bars += 1
            filled = False
            certain = True

            old_trigger = running_low * (1.0 + trail_buy_pct)
            if high >= old_trigger:
                entry_price = old_trigger
                filled = True
            else:
                new_low = low if low < running_low else running_low
                new_trigger = new_low * (1.0 + trail_buy_pct)
                if cp >= new_trigger:
                    running_low = new_low
                    entry_price = new_trigger
                    filled = True
                elif high >= new_trigger:
                    bullish = cp >= opens[i]
                    running_low = new_low
                    if bullish:
                        entry_price = new_trigger
                        filled = True
                        certain = False
                    else:
                        filled = False
                else:
                    running_low = new_low

            if filled:
                entry_certain = certain
                tp_price = entry_price * (1.0 + take_profit)
                stop_price = entry_price * (1.0 - stop_loss)
                entry_bar = i; held = 0; arm_bar = None
                in_trade = True; waiting = trailing = False
                continue
            if wait_bars >= max_hours_to_hold:
                waiting = False
            continue

        h = hours[i]
        if h != target_h0 and h != target_h1:
            continue
        di = daily_idx[i]
        if di < 0:
            continue
        sma, std = sma_arr[di], std_arr[di]
        if std == 0.0:
            continue
        lower_band = sma - std * z_thresh
        signal = (cp <= lower_band) and (cp > trend_arr[di]) if has_trend else cp <= lower_band
        if signal:
            waiting = True; running_low = cp; wait_bars = 0
            signal_bar = i; signal_z = (cp - sma) / std

    if in_trade:
        cp = prices[n - 1]
        pc = (cp - entry_price) / entry_price
        trades.append(dict(signal_i=signal_bar, signal_z=signal_z, entry_i=entry_bar,
                            entry_certain=entry_certain, arm_i=arm_bar, exit_i=n - 1,
                            entry_p=entry_price, exit_p=cp,
                            held=held, result=OPEN, ret=pc))

    return trades


def load_hourly(ticker):
    df = pd.read_csv(CACHE_DIR / f"{ticker}_1h.csv", index_col=0, parse_dates=True)
    close_col = "Adj Close" if "Adj Close" in df.columns else "Close"
    if close_col != "Close":
        df["Close"] = df[close_col]
    return df


def get_node(ticker, watchlist_id=9):
    conn = sqlite3.connect(CACHE_DIR / "trading_live.db")
    c = conn.cursor()
    c.execute(
        "SELECT window, arm_sell_pct, trail_buy_pct, trail_sell_pct, fixed_sl, "
        "max_hold_hours, z_score_threshold FROM watch_list WHERE ticker=? AND watchlist_id=?",
        (ticker, watchlist_id),
    )
    row = c.fetchone()
    conn.close()
    return dict(zip(
        ["window", "arm_sell_pct", "trail_buy_pct", "trail_sell_pct", "fixed_sl",
         "max_hold_hours", "z_score_threshold"], row))


def main(ticker):
    node = get_node(ticker)
    df_h = load_hourly(ticker)
    df_daily = df_h.resample("D").last().dropna(subset=["Close"])

    strat = strategies.TrailingBothZScoreBreakout(window=node["window"],
                                                    z_score_threshold=node["z_score_threshold"])
    ind = strat.generate_daily_indicators(df_daily)
    p = prep_inputs(df_h, ind)

    take_profit = node["arm_sell_pct"] / 100.0
    stop_loss = node["fixed_sl"] / 100.0
    trail_buy_pct = node["trail_buy_pct"] / 100.0
    trail_pct = node["trail_sell_pct"] / 100.0
    z_thresh = node["z_score_threshold"]

    raw_trades = simulate_trail_both_annotated(
        p, take_profit, stop_loss, node["max_hold_hours"],
        trail_buy_pct, trail_pct, 9, 14, z_thresh,
    )

    opens = df_h["Open"].to_numpy(dtype=float) if "Open" in df_h.columns else p['prices']
    ohlc_trades = simulate_trail_both_ohlc_aware(
        p, opens, take_profit, stop_loss, node["max_hold_hours"],
        trail_buy_pct, trail_pct, 9, 14, z_thresh,
    )

    timestamps = p['timestamps']
    sma_arr, std_arr, daily_idx = p['sma_arr'], p['std_arr'], p['daily_idx']

    def bar_sma_z(i):
        di = daily_idx[i]
        if di < 0 or std_arr[di] == 0:
            return None, None
        sma = sma_arr[di]
        z = (p['prices'][i] - sma) / std_arr[di]
        return sma, z

    trades = []
    for k, t in enumerate(raw_trades):
        pc = t['ret']
        trades.append({
            "#": k + 1,
            "Signal Time": timestamps[t['signal_i']] if t['signal_i'] is not None else None,
            "Signal Z": t['signal_z'],
            "Entry Time": timestamps[t['entry_i']],
            "Entry Price": t['entry_p'],
            "Arm Time": timestamps[t['arm_i']] if t['arm_i'] is not None else None,
            "Exit Time": timestamps[t['exit_i']],
            "Exit Price": t['exit_p'],
            "hours_held": t['held'],
            "Result": _RESULT_NAMES[t['result']],
            "Return": pc,
            "_entry_i": t['entry_i'], "_exit_i": t['exit_i'],
            "_signal_i": t['signal_i'], "_arm_i": t['arm_i'],
        })
    out = pd.DataFrame(trades)

    from openpyxl import Workbook
    from openpyxl.styles import Font

    columns = ["Trade #", "Datetime", "Open", "High", "Low", "Close", "SMA", "Z",
               "Signal Time", "Signal Z", "Entry Time", "Entry Price", "Arm Time",
               "Exit Time", "Exit Price", "hours_held", "Result", "Return", "Equity"]
    col_idx = {name: i + 1 for i, name in enumerate(columns)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(columns)
    bold = Font(bold=True)

    equity_rows = []  # row number of each trade's exit row, in order
    for _, t in out.iterrows():
        start_i = t["_signal_i"] if t["_signal_i"] is not None else t["_entry_i"]
        for i in range(start_i, t["_exit_i"] + 1):
            ts = timestamps[i]
            bar = df_h.loc[ts]
            sma, z = bar_sma_z(i)
            is_signal = i == t["_signal_i"]
            is_entry = i == t["_entry_i"]
            is_arm = i == t["_arm_i"]
            is_exit = i == t["_exit_i"]
            ws.append([
                t["#"], ts, bar.get("Open"), bar["High"], bar["Low"], bar["Close"], sma, z,
                t["Signal Time"] if is_signal else None,
                t["Signal Z"] if is_signal else None,
                t["Entry Time"] if is_entry else None,
                t["Entry Price"] if is_entry else None,
                t["Arm Time"] if is_arm else None,
                t["Exit Time"] if is_exit else None,
                t["Exit Price"] if is_exit else None,
                t["hours_held"] if is_exit else None,
                t["Result"] if is_exit else None,
                t["Return"] if is_exit else None,
                None,
            ])
            row = ws.max_row
            if is_signal or is_entry or is_arm or is_exit:
                for c in range(1, len(columns) + 1):
                    ws.cell(row=row, column=c).font = bold
            else:
                ws.row_dimensions[row].outlineLevel = 1
                ws.row_dimensions[row].hidden = True
            if is_exit:
                equity_rows.append(row)

    ws.sheet_properties.outlinePr.summaryBelow = True
    return_letter = ws.cell(row=1, column=col_idx["Return"]).column_letter
    equity_letter = ws.cell(row=1, column=col_idx["Equity"]).column_letter
    for i, row in enumerate(equity_rows):
        prev_equity = START_CAPITAL if i == 0 else f"{equity_letter}{equity_rows[i - 1]}"
        ws.cell(row=row, column=col_idx["Equity"],
                 value=f"={prev_equity}*(1+{return_letter}{row})")

    price_cols = ["Open", "High", "Low", "Close", "SMA", "Entry Price", "Exit Price"]
    for name in price_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx[name]).number_format = "0.00"
    for name in ["Z", "Signal Z"]:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx[name]).number_format = "0.00"
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col_idx["Return"]).number_format = "0.00%"
        ws.cell(row=row, column=col_idx["Equity"]).number_format = "$#,##0.00"

    compare_ws = wb.create_sheet("OHLC-Aware Compare")
    compare_cols = ["#", "Signal Time", "Entry Time (optimistic)", "Entry Price (optimistic)",
                    "Entry Time (OHLC-aware)", "Entry Price (OHLC-aware)", "Entry Shifted?",
                    "Entry Confidence", "Exit Time (optimistic)", "Exit Price (optimistic)",
                    "Result (optimistic)", "Return (optimistic)", "Exit Time (OHLC-aware)",
                    "Exit Price (OHLC-aware)", "Result (OHLC-aware)", "Return (OHLC-aware)"]
    compare_ws.append(compare_cols)
    for c in compare_ws[1]:
        c.font = bold

    n_shifted = 0
    n_heuristic = 0
    for k in range(min(len(raw_trades), len(ohlc_trades))):
        a, b = raw_trades[k], ohlc_trades[k]
        shifted = a['entry_i'] != b['entry_i']
        if shifted:
            n_shifted += 1
        if not b['entry_certain']:
            n_heuristic += 1
        compare_ws.append([
            k + 1,
            timestamps[a['signal_i']] if a['signal_i'] is not None else None,
            timestamps[a['entry_i']], a['entry_p'],
            timestamps[b['entry_i']], b['entry_p'], "YES" if shifted else "",
            "CERTAIN" if b['entry_certain'] else "HEURISTIC",
            timestamps[a['exit_i']], a['exit_p'], _RESULT_NAMES[a['result']], a['ret'],
            timestamps[b['exit_i']], b['exit_p'], _RESULT_NAMES[b['result']], b['ret'],
        ])
    if len(raw_trades) != len(ohlc_trades):
        compare_ws.append([f"Trade count differs: optimistic={len(raw_trades)}, "
                            f"OHLC-aware={len(ohlc_trades)} (a shifted entry can cascade "
                            f"into a different downstream trade sequence)"])

    price_c = {name: i + 1 for i, name in enumerate(compare_cols)}
    for name in ["Entry Price (optimistic)", "Entry Price (OHLC-aware)",
                 "Exit Price (optimistic)", "Exit Price (OHLC-aware)"]:
        for row in range(2, len(raw_trades) + 2):
            compare_ws.cell(row=row, column=price_c[name]).number_format = "0.00"
    for name in ["Return (optimistic)", "Return (OHLC-aware)"]:
        for row in range(2, len(raw_trades) + 2):
            compare_ws.cell(row=row, column=price_c[name]).number_format = "0.00%"
    for col_letter in "ABCDEFGHIJKLMNOP":
        compare_ws.column_dimensions[col_letter].width = 20

    strat_ws = wb.create_sheet("Strategy")
    strat_ws.append(["Parameter", "Value"])
    for k in strat_ws[1]:
        k.font = bold
    strat_rows = [
        ("Ticker", ticker),
        ("Strategy", "TrailingBothZScoreBreakout"),
        ("Window (days)", node["window"]),
        ("Z-Score Threshold", node["z_score_threshold"]),
        ("Trail Buy % (entry bounce)", node["trail_buy_pct"]),
        ("Arm % (take-profit arm threshold)", node["arm_sell_pct"]),
        ("Trail Sell % (trailing exit)", node["trail_sell_pct"]),
        ("Fixed Stop Loss %", node["fixed_sl"]),
        ("Max Hold Hours", node["max_hold_hours"]),
    ]
    for name, val in strat_rows:
        strat_ws.append([name, val])
    strat_ws.column_dimensions["A"].width = 34

    out_path = CACHE_DIR / f"{ticker}_trades.xlsx"
    wb.save(out_path)
    print(f"Wrote {len(out)} trades (with in-between bars) to {out_path} (starting equity ${START_CAPITAL:,})")
    n_compared = min(len(raw_trades), len(ohlc_trades))
    print(f"OHLC-aware entry timing shifted on {n_shifted}/{n_compared} trades "
          f"(trade counts: optimistic={len(raw_trades)}, OHLC-aware={len(ohlc_trades)})")
    print(f"Of OHLC-aware entries, {n_heuristic}/{len(ohlc_trades)} still rely on the "
          f"Open/Close-direction heuristic (genuinely undetermined from OHLC alone); "
          f"the rest are CERTAIN from Close-confirms-the-bounce logic.")


if __name__ == "__main__":
    ticker = sys.argv[1] if len(sys.argv) > 1 else "SOXL"
    main(ticker)
