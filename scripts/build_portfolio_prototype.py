"""One-off build script for the portfolio-construction prototype workbook
(docs/portfolio_construction_notes.md walkthrough). Not a reusable/generalized
tool -- reorders/derives columns for a specific candidate_full_review.py output
per the user's described process. Rerun by hand if the source file changes."""
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import sqlite3

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = sys.argv[1] if len(sys.argv) > 1 else "output/portfolio_prototype_20260813_205414.xlsx"
DST = sys.argv[2] if len(sys.argv) > 2 else "output/portfolio_prototype_built.xlsx"
DB = "cache/research/trading_universe.db"

wb = openpyxl.load_workbook(SRC)
ws = wb["Full Review"]
header = [c.value for c in ws[1]]
n_rows = ws.max_row

rows = []
for r in range(2, n_rows + 1):
    rows.append({header[i]: ws.cell(row=r, column=i + 1).value for i in range(len(header))})

# account_mod/account persist to candidate_nodes (2026-08-13, per user's explicit call after
# losing the original mobaxterm file) -- a manual edit already on file always wins over the
# auto-computed default, and any freshly-computed default gets written back immediately so
# nothing built here only ever lives in the xlsx.
_db_conn = sqlite3.connect(DB)
_persisted = {}
_ids = [r.get("node_id") for r in rows if r.get("node_id") is not None]
if _ids:
    placeholders = ",".join("?" * len(_ids))
    _cur = _db_conn.execute(f"SELECT id, account_mod, account FROM candidate_nodes WHERE id IN ({placeholders})", _ids)
    for cid, am, acct in _cur.fetchall():
        _persisted[cid] = (am, acct)


def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


LEVERAGE_RE = re.compile(r"^(-?\d+(?:\.\d+)?)x", re.I)


def derive(row):
    years = to_num(row.get("years"))
    trades = to_num(row.get("trades"))
    row["trades_per_year"] = round(trades / years) if trades is not None and years else None
    row["thin_flag"] = "Thin" if row["trades_per_year"] is not None and row["trades_per_year"] < 20 else ""

    for prefix in ("addon", "drought"):
        comp = to_num(row.get(f"{prefix}_compounded_pct"))
        row[f"{prefix}_cagr_pct"] = (
            round(((1 + comp / 100) ** (1 / years) - 1) * 100, 1)
            if comp is not None and years else None
        )

    note = row.get("underlier_note") or ""
    m = LEVERAGE_RE.match(note)
    row["leverage_factor"] = f"{m.group(1)}x" if m else None

    k1t = row.get("k1_tranche")
    row["tax_status"] = {
        "K1_CONFIRMED": "K1", "ETN_NOT_K1": "ETN", "CLEAN_CONFIRMED": "Clean",
    }.get(k1t, "")

    uc = to_num(row.get("underlier_count"))
    note_l = note.lower()
    if re.search(r"\bfutures\b", note_l) and not re.search(r"\bindex\b", note_l):
        structural = "Commodity"
    elif "treasury" in note_l:
        structural = "Treasury"
    elif any(w in note_l for w in ("bitcoin", "ether", "solana", "crypto")):
        structural = "Crypto"
    elif uc is not None and uc < 20:
        structural = "<20"
    else:
        structural = "Clean"
    row["structural_type"] = structural

    wn = to_num(row.get("worst_neighbor_pct"))
    if wn is None:
        row["status_3way"] = ""
    elif wn >= 0:
        row["status_3way"] = "SAFE"
    elif wn >= -20:
        row["status_3way"] = "SEMI_SAFE"
    else:
        row["status_3way"] = "CLIFF"

    cid = row.get("node_id")
    persisted_mod, persisted_account = _persisted.get(cid, (None, None))

    row["account"] = persisted_account  # manual pick -- never auto-derived

    core = to_num(row.get("strategy_cagr_pct"))
    addon = to_num(row.get("core_addon_cagr_pct"))
    drought = to_num(row.get("core_drought_cagr_pct"))
    candidates = {"Core": core, "Add On": addon, "Drought": drought}
    candidates = {k: v for k, v in candidates.items() if v is not None}
    if persisted_mod:
        row["account_mod"] = persisted_mod  # a manual edit already on file always wins
    else:
        row["account_mod"] = max(candidates, key=candidates.get) if candidates else ""
        if cid is not None and row["account_mod"]:
            _db_conn.execute("UPDATE candidate_nodes SET account_mod=? WHERE id=?", (row["account_mod"], cid))
    row["_target_cagr_sort_key"] = max(candidates.values()) if candidates else float("-inf")

    return row


for row in rows:
    derive(row)
_db_conn.commit()

# sort by target_cagr descending -- the last step, per the process: pick mods first,
# then sort on what was picked (not a raw-alpha sort before mods are even chosen)
rows.sort(key=lambda r: r["_target_cagr_sort_key"], reverse=True)

KEYS = ["node_id", "ticker", "account", "pick", "comment", "liquidity_dollars_per_day",
        "account_mod", "trades_per_year", "target_cagr", "trades",
        "years", "strategy_cagr_pct"]

GROUPS = [
    # stacked CAGR scores moved right after Keys (2026-08-13) so freeze_panes can cover
    # "Keys + core CAGR scores" as one contiguous block, without dragging along the
    # unrelated id/core/cliff/fillacc/addon/drought detail sections that used to sit
    # between them -- found live when the freeze pane ended up 50+ columns wide.
    ("stacked", ["core_addon_cagr_pct", "core_drought_cagr_pct", "core_both_cagr_pct",
                 "core_sdb_cagr_pct"]),
    ("id", ["sector", "tax_status", "structural_type", "leverage_factor", "k1_status",
             "underlier_count", "underlier_note", "candidate_type", "also_matches",
             "strategy", "liquidity_tranche"]),
    ("core", ["core_alpha_pct", "abs_return_pct", "cagr_tranche", "calendar_years_pct",
              "ann_excess_pct", "alpha_possible_pct", "alpha_pessimistic_pct",
              "alpha_certain_pct", "resolution_spread_tranche", "years_tranche",
              "trades_tranche", "thin_flag"]),
    ("cliff", ["worst_neighbor_pct", "status_3way", "status", "cliff_tranche"]),
    ("fillacc", ["fillacc_possible_win_pct", "fillacc_possible_mean_err_pct", "fillacc_n"]),
    ("addon", ["addon_n", "addon_cagr_pct", "addon_compounded_pct", "addon_win_rate_pct",
               "addon_robustness_verdict", "addon_tranche", "addon_early_wr_pct",
               "addon_late_wr_pct", "addon_wr_verdict", "addon_wr_tranche"]),
    ("drought", ["drought_n", "drought_cagr_pct", "drought_compounded_pct",
                 "drought_win_rate_pct", "drought_robustness_verdict", "drought_tranche",
                 "drought_early_wr_pct", "drought_late_wr_pct", "drought_wr_verdict",
                 "drought_wr_tranche"]),
    ("drought_ie", ["drought_ie_confirm_days", "drought_ie_vol_gate", "drought_ie_n_included",
                     "drought_ie_included_compounded_pct", "drought_ie_included_win_rate_pct",
                     "drought_ie_n_excluded", "drought_ie_excluded_compounded_pct",
                     "drought_ie_excluded_win_rate_pct", "drought_ie_verdict", "drought_ie_tranche"]),
    ("fluke", ["core_fluke_trades", "core_fluke_alpha_pct", "core_fluke_alpha_without_biggest_pct",
               "core_fluke_compounded_pct", "core_fluke_compounded_without_biggest_pct",
               "core_fluke_verdict", "core_fluke_tranche"]),
    ("wr", ["core_wr_early_pct", "core_wr_late_pct", "core_wr_diff_pct", "core_wr_verdict",
            "core_wr_tranche"]),
    ("wf", ["wf_verdict", "wf_tranche", "wf_positive_folds", "wf_total_folds",
            "wf_min_fold_alpha", "wf_max_fold_alpha", "wf_mean_fold_alpha"]),
    ("risk", ["max_drawdown_pct", "current_drawdown_pct", "trend_30d_pct", "trend_90d_pct",
              "split_flag", "drawdown_tranche", "trend_tranche", "split_tranche"]),
    ("sdb", ["sdb_trade_retention_pct", "sdb_alpha_retention_pct", "sdb_alpha_unblocked_pct",
             "sdb_alpha_blocked_pct", "sdb_compounded_unblocked_pct", "sdb_compounded_blocked_pct",
             "sdb_retention_early_pct", "sdb_retention_late_pct", "sdb_tranche", "sdb_recommend"]),
    ("cross", ["x_addon_pct", "x_drought_pct"]),
    ("exit_fillacc", ["exit_fillacc_win_pct", "exit_fillacc_mean_err_pct", "exit_fillacc_n"]),
    ("bear", ["bear_proxy", "bear_leverage", "bear_note", "bear_worst_crash",
              "bear_worst_compounded_pct", "bear_market_tranche",
              "bear_2008_gfc_decline_pct", "bear_2008_gfc_combined_pct", "bear_2008_gfc_max_dd_pct",
              "bear_2008_gfc_trades", "bear_2020_covid_decline_pct", "bear_2020_covid_combined_pct",
              "bear_2020_covid_max_dd_pct", "bear_2020_covid_trades", "bear_2022_bear_decline_pct",
              "bear_2022_bear_combined_pct", "bear_2022_bear_max_dd_pct", "bear_2022_bear_trades",
              "bear_2000_dotcom_decline_pct", "bear_2000_dotcom_combined_pct",
              "bear_2000_dotcom_max_dd_pct", "bear_2000_dotcom_trades", "bear_2000_dotcom_truncated"]),
    ("crash25", ["crash25_ticker_bh_decline_pct", "crash25_ticker_bh_recovery_pct",
                 "crash25_ticker_bh_combined_pct", "crash25_ticker_bh_max_dd_pct",
                 "crash25_algo_decline_pct", "crash25_algo_recovery_pct",
                 "crash25_algo_combined_pct", "crash25_algo_max_dd_pct", "crash25_algo_trades",
                 "crash25_verdict", "crash_2025_tranche"]),
]

final_cols = list(KEYS)
group_bounds = []  # (group_name, start_col_idx_1based, end_col_idx_1based)
for name, cols in GROUPS:
    start = len(final_cols) + 1
    final_cols.extend(cols)
    end = len(final_cols)
    group_bounds.append((name, start, end))

missing = [c for c in final_cols if c not in ("account_mod", "target_cagr") and c not in rows[0]]
if missing:
    print("WARNING missing columns:", missing)

col_idx = {name: i + 1 for i, name in enumerate(final_cols)}

new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = "Full Review"
DISPLAY_NAME = {"node_id": "c_id"}  # c_id = candidate_nodes.id, distinct from live watch_list.id
for i, name in enumerate(final_cols, 1):
    new_ws.cell(row=1, column=i, value=DISPLAY_NAME.get(name, name))

for r, row in enumerate(rows, start=2):
    for name, ci in col_idx.items():
        if name == "target_cagr":
            continue
        new_ws.cell(row=r, column=ci, value=row.get(name))

# target_cagr formula: IF(account_mod, map to the matching CAGR cell). account_mod is
# auto-picked above (best of Core/Add On/Drought, Overlay excluded per user's call) but
# still a real cell value, so a manual override just means typing over it.
mod_col = get_column_letter(col_idx["account_mod"])
core_col = get_column_letter(col_idx["strategy_cagr_pct"])
addon_col = get_column_letter(col_idx["core_addon_cagr_pct"])
drought_col = get_column_letter(col_idx["core_drought_cagr_pct"])
overlay_col = get_column_letter(col_idx["core_both_cagr_pct"])
target_col_letter = get_column_letter(col_idx["target_cagr"])
max_of_three = (f'MAX({core_col}{{r}},{addon_col}{{r}},{drought_col}{{r}})')
for r in range(2, n_rows + 1):
    fallback = max_of_three.format(r=r)
    formula = (
        f'=IF({mod_col}{r}="Core",{core_col}{r},'
        f'IF({mod_col}{r}="Add On",{addon_col}{r},'
        f'IF({mod_col}{r}="Drought",{drought_col}{r},'
        f'IF({mod_col}{r}="Overlay",{overlay_col}{r},{fallback}))))'
    )
    cell = new_ws.cell(row=r, column=col_idx["target_cagr"], value=formula)
    cell.number_format = '#,##0"%"'

# data validation dropdown for account_mod
dv = DataValidation(type="list", formula1='"Core,Add On,Drought,Overlay"', allow_blank=True)
new_ws.add_data_validation(dv)
dv.add(f"{mod_col}2:{mod_col}{n_rows}")

# data validation dropdown for account (manual pick, not derived)
account_col_letter = get_column_letter(col_idx["account"])
dv2 = DataValidation(type="list", formula1='"brokerage,ira,roth,soxl_ira"', allow_blank=True)
new_ws.add_data_validation(dv2)
dv2.add(f"{account_col_letter}2:{account_col_letter}{n_rows}")

# number formats
pct_cols = [c for c in final_cols if c.endswith("_pct")]
for name in pct_cols:
    ci = col_idx[name]
    letter = get_column_letter(ci)
    for r in range(2, n_rows + 1):
        new_ws.cell(row=r, column=ci).number_format = '#,##0"%"'
liq_letter = get_column_letter(col_idx["liquidity_dollars_per_day"])
for r in range(2, n_rows + 1):
    new_ws.cell(row=r, column=col_idx["liquidity_dollars_per_day"]).number_format = "#,##0"

# freeze panes through the core CAGR score block (strategy_cagr_pct is already in Keys;
# this extends the frozen area through the stacked core_addon/drought/both/sdb columns
# too, so the whole CAGR picture stays visible while scrolling right)
freeze_after = col_idx["core_sdb_cagr_pct"]
freeze_col = get_column_letter(freeze_after + 1)
new_ws.freeze_panes = f"{freeze_col}2"
new_ws.auto_filter.ref = new_ws.dimensions
for i, name in enumerate(final_cols, 1):
    letter = get_column_letter(i)
    width = 10 if name in ("comment", "underlier_note", "bear_note") else min(8, max(5, len(name) // 2))
    new_ws.column_dimensions[letter].width = width

# single column group: worst_neighbor_pct (cliff-safety) through drought_wr_tranche
# (cliff+fillacc+addon+drought detail), collapsed by default.
# openpyxl's column_dimensions.group() only sets outlineLevel/hidden on the FIRST
# column of the range, not the whole span (confirmed via a minimal repro 2026-08-13) --
# looping explicitly instead of relying on that convenience method.
group_start = col_idx["worst_neighbor_pct"]
group_end = col_idx["drought_wr_tranche"]
for i in range(group_start, group_end + 1):
    dim = new_ws.column_dimensions[get_column_letter(i)]
    dim.outlineLevel = 1
    dim.hidden = True
new_ws.sheet_properties.outlinePr.summaryRight = False

# conditional highlighting: addon CAGR cell highlighted if addon_tranche == OK (same row)
GREEN = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_RED = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
addon_cagr_letter = get_column_letter(col_idx["core_addon_cagr_pct"])
addon_tranche_letter = get_column_letter(col_idx["addon_tranche"])
drought_cagr_letter = get_column_letter(col_idx["core_drought_cagr_pct"])
drought_tranche_letter = get_column_letter(col_idx["drought_tranche"])
overlay_cagr_letter = get_column_letter(col_idx["core_both_cagr_pct"])
core_cagr_letter = get_column_letter(col_idx["strategy_cagr_pct"])
data_range = f"2:{n_rows}"

# strategy_cagr_pct: static light-red base fill (every row)
for r in range(2, n_rows + 1):
    new_ws.cell(row=r, column=col_idx["strategy_cagr_pct"]).fill = LIGHT_RED

# yellow (which CAGR column account_mod actually picked) takes priority over red/green --
# added first + stopIfTrue so it wins when both would otherwise apply to the same cell
new_ws.conditional_formatting.add(
    f"{core_cagr_letter}2:{core_cagr_letter}{n_rows}",
    FormulaRule(formula=[f'${mod_col}2="Core"'], fill=YELLOW, stopIfTrue=True)
)
new_ws.conditional_formatting.add(
    f"{addon_cagr_letter}2:{addon_cagr_letter}{n_rows}",
    FormulaRule(formula=[f'${mod_col}2="Add On"'], fill=YELLOW, stopIfTrue=True)
)
new_ws.conditional_formatting.add(
    f"{drought_cagr_letter}2:{drought_cagr_letter}{n_rows}",
    FormulaRule(formula=[f'${mod_col}2="Drought"'], fill=YELLOW, stopIfTrue=True)
)

new_ws.conditional_formatting.add(
    f"{addon_cagr_letter}2:{addon_cagr_letter}{n_rows}",
    FormulaRule(formula=[f'${addon_tranche_letter}2="OK"'], fill=GREEN)
)
new_ws.conditional_formatting.add(
    f"{drought_cagr_letter}2:{drought_cagr_letter}{n_rows}",
    FormulaRule(formula=[f'${drought_tranche_letter}2="OK"'], fill=GREEN)
)
new_ws.conditional_formatting.add(
    f"{overlay_cagr_letter}2:{overlay_cagr_letter}{n_rows}",
    FormulaRule(formula=[f'AND(${addon_tranche_letter}2="OK",${drought_tranche_letter}2="OK")'], fill=GREEN)
)

# Column Definitions sheet, carried over + new columns appended
if "Column Definitions" in wb.sheetnames:
    src_def = wb["Column Definitions"]
    new_def = new_wb.create_sheet("Column Definitions")
    for r in src_def.iter_rows(values_only=True):
        if r and r[0] == "node_id":
            r = ("c_id", r[1] + " Renamed from node_id for clarity -- candidate_nodes.id, "
                                 "NOT the live watch_list.id used by the trading daemon.")
        new_def.append(r)
    NEW_DEFS = {
        "account_mod": "Manual pick: Core / Add On / Drought / Overlay. Drives target_cagr via formula.",
        "trades_per_year": "trades / years, derived.",
        "target_cagr": "Formula: looks up the CAGR cell matching this row's account_mod pick.",
        "leverage_factor": "Parsed from underlier_note's leading pattern (e.g. -2x/3x/1.5x).",
        "tax_status": "K1 / ETN / Clean, derived from k1_tranche (informational/account-routing, not disqualifying).",
        "structural_type": "Commodity / Treasury / Crypto / <20 / Clean, derived from underlier_note + underlier_count. Priority: Commodity > Treasury > Crypto > <20 > Clean.",
        "status_3way": "SAFE (worst_neighbor_pct>=0) / SEMI_SAFE (-20<=x<0) / CLIFF (<-20). Looser than the binary status/cliff_tranche.",
        "thin_flag": '"Thin" if trades_per_year < 20 (most likely unreliable).',
        "addon_cagr_pct": "addon_compounded_pct annualized via years: (1+compounded/100)^(1/years)-1.",
        "drought_cagr_pct": "drought_compounded_pct annualized via years, same formula as addon_cagr_pct.",
    }
    for name, definition in NEW_DEFS.items():
        new_def.append((name, definition))

new_wb.save(DST)
print(f"Wrote {DST} -- {len(final_cols)} columns, {n_rows-1} rows")
